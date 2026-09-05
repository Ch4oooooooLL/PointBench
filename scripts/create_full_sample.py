"""生成 PointProcess 全要素演示数据包（可直接完整导入）。

产出（默认写入 ``sample_data/``）：

- ``POINTPROCESS_DEMO_FULL_<date>.zip``            完整备份导入包
- ``POINTPROCESS_DEMO_FULL_<date>_summary.json``   包内容摘要
- ``POINTPROCESS_DEMO_COMPLEX_FEM_REPLACEMENT_<date>.zip``  复杂 FEM 替换包（--mode fem-replacement）

包内同时携带 ``manifest.json``（应用兼容视图）与 ``pointprocess_backup.json``
（完整迁移数据），并包含 ``fem/`` 源文件目录（主文件 + INCLUDE 子文件）。
通过 导入 → 上传 zip → 预览 → 确认导入 即可一次性恢复：

- 项目（车架疲劳台架试验）
- 8 个点位：通道、CAE 映射、安装/检查状态、点位照片
- 10 个测试轮次 × 8 点位测量数据（含异常记录）
- 12 条裂缝记录（裂缝照片，关联点位与轮次）
- 2 条 Dewesoft 导入记录（含匹配/未匹配通道）
- FEM 模型（OptiStruct 格式，导入端自动解析并生成 GLB 渲染产物）

用法::

    python scripts/create_full_sample.py [--out sample_data] [--seed 20260905]
    python scripts/create_full_sample.py --mode fem-replacement   # 复杂 FEM 替换包

``--mode full``（默认）把第一版骨架 FEM（``frame_assembly_simple.fem``）装入完整
演示包，用于先导入项目；``--mode fem-replacement`` 单独生成第二版高保真复杂 FEM
（``frame_assembly_v2.fem``，更多节点/单元/部件与 CQUAD4+CTRIA3+CROD 混合单元），
该文件为自包含单文件（无 INCLUDE 依赖），可在 FEM 预览页直接选择上传，把项目内
已渲染的骨架模型替换为复杂模型并重新渲染。

脚本在打包前会用后端 FEM 解析器实际解析生成的模型文件，确保随包分发的
FEM 可被系统直接解析。
"""

from __future__ import annotations

import argparse
import json
import math
import random
import sys
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT / "backend"
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

PROJECT_ID = "POINTPROCESS-DEMO-FULL-20260905"
PROJECT_NAME = "车架疲劳台架试验 — 全要素演示项目"
EXPORT_ID = "DEMO-FULL-20260905-FULL-BACKUP"

TOTAL_RUNS = 10
CYCLE_STEP = 8000
ABNORMAL_AMPLITUDE_UE = 1300.0

CATEGORY_META = {
    "A": {"label": "稳定型", "desc": "应变基本不变，仅微量噪声"},
    "B": {"label": "规律变动型", "desc": "线性增长叠加正弦规律"},
    "C": {"label": "复杂变动型", "desc": "涨跌交替，偶发突变"},
    "D": {"label": "剧烈变动型", "desc": "大幅跳变，尖峰骤降"},
}

# 点位布局与 FEM 模型部件一一对应，便于对照"测点 ↔ CAE 模型"。
POINT_PLAN = [
    ("01", "A", "左纵梁", "left", "principal", "low"),
    ("02", "A", "右纵梁", "right", "principal", "low"),
    ("03", "B", "左纵梁", "left", "longitudinal", "medium"),
    ("04", "B", "右纵梁", "right", "longitudinal", "medium"),
    ("05", "C", "前横梁", "front", "transverse", "high"),
    ("06", "C", "中部横梁", "middle", "transverse", "high"),
    ("07", "D", "后横梁", "rear", "transverse", "critical"),
    ("08", "D", "焊缝加强板", "middle", "principal", "critical"),
]

TAGS_MAP = {
    "A": ["稳定区", "非关键点"],
    "B": ["规律变化", "疲劳监测"],
    "C": ["复杂变化", "CAE对应点", "需关注"],
    "D": ["剧烈变动", "危险点", "CAE对应点", "重点监控"],
}

CATEGORY_COLORS = {"A": "#dce7ea", "B": "#d4edda", "C": "#fff3cd", "D": "#f8d7da"}


# ═══════════════════════════════════════════════════════════════
#  SVG 图片生成（点位照片 / 裂缝照片）
# ═══════════════════════════════════════════════════════════════

def photo_svg(point_id: str, photo_type: str, category: str, component: str) -> bytes:
    color = CATEGORY_COLORS[category]
    label = "总览" if photo_type == "overview" else "细节"
    return f"""<svg xmlns="http://www.w3.org/2000/svg" width="800" height="560">
<rect width="800" height="560" fill="{color}"/>
<rect x="60" y="80" width="680" height="360" rx="12" fill="#ffffff" opacity="0.9"/>
<line x1="60" y1="170" x2="740" y2="170" stroke="#9ab5bd" stroke-dasharray="6 4"/>
<line x1="60" y1="260" x2="740" y2="260" stroke="#9ab5bd" stroke-dasharray="6 4"/>
<line x1="60" y1="350" x2="740" y2="350" stroke="#9ab5bd" stroke-dasharray="6 4"/>
<circle cx="400" cy="260" r="64" fill="none" stroke="#c2504a" stroke-width="5"/>
<circle cx="400" cy="260" r="10" fill="#c2504a"/>
<text x="400" y="140" font-family="Arial" font-size="44" text-anchor="middle" fill="#172026">测点 {point_id} {label}照片</text>
<text x="400" y="480" font-family="Arial" font-size="26" text-anchor="middle" fill="#5f6f76">{component} · 类别 {category} · 演示数据</text>
</svg>""".encode("utf-8")


def crack_svg(point_id: str, cycle_count: int, seq: int) -> bytes:
    """生成带锯齿裂缝形态的示意图（随循环次数增加裂缝变长）。"""
    length = min(300, 120 + cycle_count // CYCLE_STEP * 20 + seq * 10)
    x0, y0 = 180, 300
    segments: list[str] = []
    random.seed(f"crack-{point_id}-{cycle_count}")
    x, y = x0, y0
    steps = 8
    for index in range(steps):
        x += length / steps
        y += random.randint(-26, 26)
        segments.append(f'L {x:.0f} {y:.0f}')
    path = f"M {x0} {y0} " + " ".join(segments)
    return f"""<svg xmlns="http://www.w3.org/2000/svg" width="800" height="560">
<rect width="800" height="560" fill="#e8e2d6"/>
<rect x="60" y="80" width="680" height="400" rx="10" fill="#cfd6da"/>
<line x1="60" y1="280" x2="740" y2="280" stroke="#aab6bc" stroke-width="2"/>
<line x1="120" y1="80" x2="120" y2="480" stroke="#aab6bc" stroke-width="2"/>
<path d="{path}" stroke="#7a1f1f" stroke-width="6" fill="none" stroke-linecap="round"/>
<circle cx="{x0}" cy="{y0}" r="10" fill="#7a1f1f"/>
<text x="400" y="60" font-family="Arial" font-size="34" text-anchor="middle" fill="#172026">测点 {point_id} 裂缝照片 — 循环 {cycle_count} 次</text>
<text x="400" y="530" font-family="Arial" font-size="22" text-anchor="middle" fill="#5f6f76">裂缝扩展观测（演示数据）</text>
</svg>""".encode("utf-8")


# ═══════════════════════════════════════════════════════════════
#  FEM 模型生成（OptiStruct/Nastran 格式，主文件 + INCLUDE 子文件）
# ═══════════════════════════════════════════════════════════════

def _card(name: str, *fields: object) -> str:
    return name.ljust(8) + "".join(str(field).rjust(8) for field in fields)


def _fmt(value: float) -> str:
    return f"{value:.1f}"


class _MeshBuilder:
    """FEM 板壳网格生成器。

    所有 id 全局单调递增；``begin_section`` 起记录当前部件/子文件的
    grid 与 element 行。CQUAD4 使用常规 8 字段格式（解析器同样支持紧凑格式）。
    """

    def __init__(self) -> None:
        self._next_node = 1
        self._next_elem = 1
        self._grids: list[str] = []
        self._elems: list[str] = []

    def begin_section(self) -> None:
        self._grids = []
        self._elems = []

    def grid(self, x: float, y: float, z: float) -> int:
        nid = self._next_node
        self._next_node += 1
        self._grids.append(_card("GRID", nid, "", _fmt(x), _fmt(y), _fmt(z)))
        return nid

    def quad_cards(self, pid: int, ids: list[int]) -> None:
        eid = self._next_elem
        self._next_elem += 1
        self._elems.append(_card("CQUAD4", eid, pid, *ids))

    def quad(self, pid: int, p0, p1, p2, p3) -> None:
        nids = [self.grid(*pp) for pp in (p0, p1, p2, p3)]
        self.quad_cards(pid, nids)

    def tria(self, pid: int, p0, p1, p2) -> None:
        nids = [self.grid(*pp) for pp in (p0, p1, p2)]
        self.tria_cards(pid, nids)

    def tria_cards(self, pid: int, ids: list[int]) -> None:
        eid = self._next_elem
        self._next_elem += 1
        self._elems.append(_card("CTRIA3", eid, pid, *ids))

    def ring(
        self,
        stations: list[float],
        *,
        center: tuple[float, float],
        rx: float,
        ry: float,
        pid: int,
    ) -> None:
        """沿 z 轴扫掠矩形截面，每相邻两站生成 4 个 CQUAD4 面。"""
        cx, cy = center
        rect = [(-rx, -ry), (rx, -ry), (rx, ry), (-rx, ry)]
        station_nodes: list[list[int]] = []
        for z in stations:
            station_nodes.append([self.grid(cx + dx, cy + dy, z) for dx, dy in rect])
        for index in range(len(stations) - 1):
            a, b = station_nodes[index], station_nodes[index + 1]
            self.quad_cards(pid, [a[0], a[1], b[1], b[0]])
            self.quad_cards(pid, [a[1], a[2], b[2], b[1]])
            self.quad_cards(pid, [a[2], a[3], b[3], b[2]])
            self.quad_cards(pid, [a[3], a[0], b[0], b[3]])

    def grids(self) -> list[str]:
        return self._grids

    def elems(self) -> list[str]:
        return self._elems


def _linspace_stations(start: float, end: float, spacing: float) -> list[float]:
    count = max(2, round((end - start) / spacing) + 1)
    return [start + (end - start) * index / (count - 1) for index in range(count)]


def _component(name: str, comp_id: int, hw_color: int, lines: list[str]) -> list[str]:
    return [
        f"$$ ---- {name} ----",
        f'$HMNAME COMP {comp_id} "{name}"',
        f"$HWCOLOR COMP {comp_id} {hw_color}",
        f"$HMCOMP ID {comp_id}",
        *lines,
        "$",
    ]


def build_simple_fem_source_files() -> dict[str, str]:
    """第一版演示 FEM：骨架级车架模型（全部 CQUAD4）。

    - 主文件 frame_assembly_simple.fem + INCLUDE parts/rails_simple.fem
    - 左右纵梁为矩形截面扫掠管，3 根横梁 + 4 块焊缝加强板，共 6 个部件
    - 用于先导入项目，随后在项目内用第二版（复杂）模型整体替换
    """
    builder = _MeshBuilder()

    def _rail_include() -> str:
        sections: list[str] = []
        for side_name, y_center, comp_id, hw_color, pid in (
            ("左纵梁", 160.0, 1, 5, 1),
            ("右纵梁", -160.0, 2, 1, 2),
        ):
            builder.begin_section()
            builder.ring(
                _linspace_stations(-600.0, 600.0, 150.0),
                center=(0.0, y_center),
                rx=60.0,
                ry=80.0,
                pid=pid,
            )
            sections.extend(
                [
                    f'$HMNAME COMP {comp_id} "{side_name}"',
                    f"$HWCOLOR COMP {comp_id} {hw_color}",
                    f"$HMCOMP ID {comp_id}",
                    *builder.grids(),
                    *builder.elems(),
                    "",
                ]
            )
        return "\n".join(["$$ 左右纵梁子文件（第一版）", *sections])

    main_sections: list[str] = [
        "$$ =================================================",
        "$$ PointProcess 车架演示 - 第一版骨架模型（用于替换演示）",
        "$$ 单位: mm / N / MPa   生成: scripts/create_full_sample.py",
        "$$ =================================================",
        "BEGIN BULK",
        "$",
        "$$ ---- 材料: 普通钢材 ----",
        _card("MAT1", 1, "2.06+5", "7.91+4", ".3", "7.85-9"),
        "$",
        "$$ ---- 属性 ----",
        _card("PSHELL", 1, 1, "3.0"),
        _card("PSHELL", 2, 1, "3.0"),
        _card("PSHELL", 3, 1, "4.0"),
        _card("PSHELL", 4, 1, "4.0"),
        _card("PSHELL", 5, 1, "4.0"),
        _card("PSHELL", 6, 1, "5.0"),
        "$",
        "$$ ---- 纵梁（INCLUDE 子文件，部件 1/2）----",
        "INCLUDE parts/rails_simple.fem",
        "$",
    ]
    for name, comp_id, hw_color, pid, x_center in (
        ("前横梁", 3, 3, 3, -300.0),
        ("中部横梁", 4, 4, 4, 0.0),
        ("后横梁", 5, 2, 5, 300.0),
    ):
        builder.begin_section()
        builder.ring(
            [0.0, 60.0, 130.0, 200.0, 260.0],
            center=(x_center, 0.0),
            rx=50.0,
            ry=60.0,
            pid=pid,
        )
        main_sections.extend(_component(name, comp_id, hw_color, [*builder.grids(), *builder.elems()]))
    # 焊缝加强板：前/后横梁与纵梁交接处 4 块竖直板
    builder.begin_section()
    for x_center in (-300.0, 300.0):
        for y_center in (160.0, -160.0):
            builder.quad(
                6,
                (x_center, y_center - 30.0, 0.0),
                (x_center, y_center + 30.0, 0.0),
                (x_center, y_center + 30.0, 80.0),
                (x_center, y_center - 30.0, 80.0),
            )
    main_sections.extend(_component("焊缝加强板", 6, 6, [*builder.grids(), *builder.elems()]))
    main_sections.extend(["", "ENDDATA", ""])
    return {
        "frame_assembly_simple.fem": "\n".join(main_sections),
        "parts/rails_simple.fem": _rail_include(),
    }


def build_complex_fem_source_files(single_file: bool = False) -> dict[str, str]:
    """第二版演示 FEM：高保真车架模型（更复杂，用于整体替换演示）。

    与第一版对比：
    - 部件更多：左右纵梁（盖板+腹板多 PID）、8 根等距横梁、前后保险杠、
      四角立柱与保险杠连杆，共 14 个部件
    - 单元类型更丰富：CQUAD4 壳 + CTRIA3 过渡 + CROD 梁（圆形截面属性）
    - 纵梁为闭口箱型截面（上/下盖板 + 内/外腹板），网格更细
    - 横梁 8 根等距布置，PID 随序号交替（5.0/6.0 板厚）
    - 立柱顶用 CTRIA3 封口，连杆用 CROD

    ``single_file=True`` 时把纵梁（子文件内容）直接内联进主文件，返回
    ``{"frame_assembly_v2.fem": ...}`` 单一自包含模型，无需随附子文件。
    """
    builder = _MeshBuilder()

    def _rail_deck() -> str:
        sections: list[str] = []
        for side_name, y_center, comp_id, hw_color in (
            ("左纵梁", 160.0, 1, 5),
            ("右纵梁", -160.0, 2, 1),
        ):
            builder.begin_section()
            # 闭口箱型截面：每站 4 个角点（底/顶 × 内/外），相邻站间直接构面。
            # 上/下盖板 PID=1（板厚 3.0），内/外腹板 PID=3（板厚 4.0）。
            prev: list[int] | None = None
            for x in _linspace_stations(-600.0, 600.0, 60.0):
                n_bl = builder.grid(x, y_center - 30.0, 0.0)
                n_br = builder.grid(x, y_center + 30.0, 0.0)
                n_tl = builder.grid(x, y_center - 30.0, 80.0)
                n_tr = builder.grid(x, y_center + 30.0, 80.0)
                if prev is not None:
                    p_bl, p_br, p_tl, p_tr = prev
                    builder.quad_cards(1, [p_bl, p_br, n_br, n_bl])  # 下盖板
                    builder.quad_cards(1, [p_tl, p_tr, n_tr, n_tl])  # 上盖板
                    builder.quad_cards(3, [p_bl, p_tl, n_tl, n_bl])  # 外腹板
                    builder.quad_cards(3, [p_br, p_tr, n_tr, n_br])  # 内腹板
                prev = [n_bl, n_br, n_tl, n_tr]
            sections.extend(
                [
                    f'$HMNAME COMP {comp_id} "{side_name}"',
                    f"$HWCOLOR COMP {comp_id} {hw_color}",
                    f"$HMCOMP ID {comp_id}",
                    *builder.grids(),
                    *builder.elems(),
                    "",
                ]
            )
        return "\n".join(["$$ 左右纵梁（箱型截面，盖板+腹板多 PID）", *sections])

    mats = [
        _card("MAT1", 1, "2.06+5", "7.91+4", ".3", "7.85-9"),
        _card("MAT1", 2, "7.0+4", "2.7+4", ".33", "2.7-9"),
    ]
    props = [
        _card("PSHELL", 1, 1, "3.0"),
        _card("PSHELL", 2, 1, "3.0"),
        _card("PSHELL", 3, 1, "4.0"),
        _card("PSHELL", 4, 1, "5.0"),
        _card("PSHELL", 5, 1, "6.0"),
        _card("PSHELL", 6, 1, "8.0"),
        _card("PSHELL", 7, 2, "3.0"),
        _card("PBARL", 11, 1, "ROD"),
        "+       " + "".join(str(v).rjust(8) for v in (2.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0)),
        "+       " + "".join(str(v).rjust(8) for v in (0.0, 0.0, 0.0, 0.0, 0.0, 0.0)),
    ]

    def _cross_beam(x_center: float, comp_id: int, hw_color: int, pid: int) -> list[str]:
        builder.begin_section()
        builder.ring(
            [0.0, 40.0, 100.0, 160.0, 220.0, 260.0],
            center=(x_center, 0.0),
            rx=45.0,
            ry=55.0,
            pid=pid,
        )
        return _component(f"横梁{comp_id - 10}", comp_id, hw_color, [*builder.grids(), *builder.elems()])

    def _bumper(x_sign: float, comp_id: int, hw_color: int, name: str) -> list[str]:
        builder.begin_section()
        builder.ring(
            [-180.0, -80.0, 0.0, 80.0, 180.0],
            center=(x_sign * 620.0, 0.0),
            rx=40.0,
            ry=45.0,
            pid=6,
        )
        return _component(name, comp_id, hw_color, [*builder.grids(), *builder.elems()])

    def _pillars_and_rods() -> list[str]:
        builder.begin_section()
        pillar_top: list[int] = []
        rod_lines: list[str] = []
        for x_sign in (1.0, -1.0):
            for y_sign in (1.0, -1.0):
                y0, y1 = y_sign * 120.0, y_sign * 160.0
                x0 = x_sign * 600.0
                # 立柱四面（朝向保险杠的面）+ 顶部 CTRIA3 封口
                n_bot = [
                    builder.grid(x0, y0, 80.0),
                    builder.grid(x0, y1, 80.0),
                    builder.grid(x0, y1, 190.0),
                    builder.grid(x0, y0, 190.0),
                ]
                builder.quad_cards(6, [n_bot[0], n_bot[1], n_bot[2], n_bot[3]])
                n_top = [
                    builder.grid(x0 + 10.0, y0 + 10.0, 190.0),
                    builder.grid(x0 + 10.0, y1 - 10.0, 190.0),
                ]
                # 顶部横向板 + 斜封口（CTRIA3，基于既有节点）
                builder.quad_cards(6, [n_bot[3], n_bot[2], n_top[1], n_top[0]])
                builder.tria_cards(6, [n_bot[0], n_top[0], n_top[1]])
                builder.tria_cards(6, [n_bot[0], n_top[1], n_bot[1]])
                pillar_top.append(n_top[0])
                pillar_top.append(n_top[1])
        # 保险杠连杆：立柱顶部 → 保险杠端部上缘（CROD）
        rod_targets = [
            (x_sign * 660.0, y_sign * 120.0, 160.0)
            for x_sign in (1.0, -1.0)
            for y_sign in (1.0, -1.0)
        ]
        for index, target in enumerate(rod_targets):
            g2 = builder.grid(*target)
            eid = builder._next_elem
            builder._next_elem += 1
            rod_lines.append(_card("CROD", eid, 11, pillar_top[index], g2))
        return _component("四角立柱与连杆", 9, 3, [*builder.grids(), *builder.elems(), *rod_lines])

    # 纵梁网格先生成（id 从 1 起、文本靠前），无论内联还是 INCLUDE 都保持连续。
    rails_text = _rail_deck()

    sections: list[str] = [
        "$$ =================================================",
        "$$ PointProcess 车架演示 - 第二版高保真模型（用于整体替换）",
        "$$ 单元: CQUAD4 壳 + CTRIA3 过渡 + CROD 梁",
        "$$ 部件: 左右纵梁 / 8 根横梁 / 前后保险杠 / 立柱连杆",
        "$$ 单位: mm / N / MPa   生成: scripts/create_full_sample.py",
        "$$ =================================================",
        "BEGIN BULK",
        "$",
        "$$ ---- 材料 ----",
        *mats,
        "$",
        "$$ ---- 属性 ----",
        *props,
        "$",
        "$$ ---- 左右纵梁 ----",
    ]
    if single_file:
        sections.append(rails_text)
    else:
        sections.append("INCLUDE parts/rails_v2.fem")
        sections.append("$")
    for index in range(1, 9):
        x_center = -560.0 + (index - 1) * 160.0
        sections.extend(_cross_beam(x_center, 10 + index, 3 if index <= 4 else 2, 4 if index % 2 else 5))
    sections.extend(_bumper(1.0, 7, 5, "前保险杠"))
    sections.extend(_bumper(-1.0, 8, 6, "后保险杠"))
    sections.extend(_pillars_and_rods())
    sections.extend(["", "ENDDATA", ""])
    if single_file:
        return {"frame_assembly_v2.fem": "\n".join(sections)}
    return {
        "frame_assembly_v2.fem": "\n".join(sections),
        "parts/rails_v2.fem": rails_text,
    }


def validate_fem_source(
    source_dir: Path,
    main_filename: str = "frame_assembly.fem",
    include_count: int = 1,
) -> dict[str, int]:
    """用后端解析器实际解析生成的 FEM，确保随包模型可被系统导入。

    ``include_count`` 期望被解析器加载的 INCLUDE 子文件数量：主文件 + 子文件
    模型传 1；自包含单文件模型传 0。
    """
    from app.services.fem.parser import FemModelProvider

    main_path = source_dir / main_filename
    model = FemModelProvider(main_path, include_root=source_dir).load()
    stats = {
        "node_count": len(model.nodes),
        "element_count": len(model.elements),
        "component_count": len(model.components),
        "included_files": len(model.metadata.get("included_files", [])),
    }
    if stats["node_count"] == 0 or stats["element_count"] == 0:
        raise RuntimeError(f"FEM 校验失败：模型为空 {stats}")
    if not model.metadata.get("element_component_ids"):
        raise RuntimeError("FEM 校验失败：部件分组信息缺失（$HMCOMP 块未被识别）")
    if stats["included_files"] != include_count:
        raise RuntimeError(f"FEM 校验失败：INCLUDE 子文件数量不符，期望 {include_count}，实际 {stats['included_files']}")
    return stats


# ═══════════════════════════════════════════════════════════════
#  测量数据生成（四类行为模式）
# ═══════════════════════════════════════════════════════════════

def _strain_for(category: str, point_seq: int, run_index: int, total_runs: int) -> tuple[float, float]:
    progress = run_index / total_runs
    base_amplitude = 120.0 + point_seq * 90.0
    if category == "A":
        base = 85.0 + point_seq * 11.0
        noise = random.gauss(0, 1.5)
        return round(base + noise, 2), round(base - 6.0 + noise, 2)
    if category == "B":
        amplitude = base_amplitude * (1.0 + progress * 0.35)
        amplitude += math.sin(progress * math.pi * 2.5) * base_amplitude * 0.06
        mean = 10.0 + point_seq * 2.0 + run_index * 2.0
        return round(mean + amplitude, 2), round(mean - amplitude, 2)
    if category == "C":
        trend = 1.0 + 0.3 * math.sin(progress * math.pi * 1.8)
        mid = 0.12 * math.sin(progress * math.pi * 4.5)
        spike = 0.0
        if run_index == 3:
            spike = base_amplitude * 0.28
        elif run_index == 7:
            spike = -base_amplitude * 0.22
        amplitude = base_amplitude * (trend + mid) + spike
        mean = 10.0 + point_seq * 2.0 + run_index * 0.8
        return round(mean + abs(amplitude), 2), round(mean - abs(amplitude), 2)
    # D: 剧烈变动
    amp_mult = 1.0 + 0.6 * math.sin(progress * math.pi * 1.3)
    spikes = {2: 0.55, 5: 0.70, 8: -0.45, 9: 0.85}
    spike = base_amplitude * spikes.get(run_index, 0.0)
    jitter = random.gauss(0, base_amplitude * 0.08)
    amplitude = max(base_amplitude * amp_mult + spike + jitter, base_amplitude * 0.3)
    mean = 10.0 + point_seq * 3.0 + run_index * 0.5 + 12.0 * math.sin(progress * math.pi * 3.0)
    return round(mean + abs(amplitude), 2), round(mean - abs(amplitude), 2)


# ═══════════════════════════════════════════════════════════════
#  数据组装
# ═══════════════════════════════════════════════════════════════

def build_dataset(now: str, start_time: datetime) -> dict[str, Any]:
    random.seed(20260905)
    points: list[dict[str, Any]] = []
    package_files: dict[str, bytes] = {}

    for point_seq, (point_id, category, component, side, direction, danger) in enumerate(POINT_PLAN, start=1):
        meta = CATEGORY_META[category]
        photos = []
        for photo_seq, photo_type in enumerate(("overview", "detail"), start=1):
            photo_path = f"photos/{point_id}/{point_id}_{photo_type}_001.svg"
            package_files[photo_path] = photo_svg(point_id, photo_type, category, component)
            photos.append(
                {
                    "photo_id": f"P-{point_id}-00{photo_seq}",
                    "type": photo_type,
                    "path": photo_path,
                    "filename": Path(photo_path).name,
                    "taken_time": now,
                    "sha256": "",
                    "remark": "总览图" if photo_type == "overview" else "细节图",
                }
            )
        points.append(
            {
                "point_id": point_id,
                "point_name": f"车架疲劳测点 {point_id}（{meta['label']}）",
                "point_type": "strain_gauge",
                "component": component,
                "side": side,
                "position_description": f"{component} — {meta['desc']}，{side} 侧 {direction} 方向贴片",
                "direction": direction,
                "bridge_type": "1/4_bridge" if point_seq % 2 else "1/2_bridge",
                "resistance_ohm": round(120.0 + point_seq * 0.08, 2),
                "install_status": "installed",
                "check_status": "checked",
                "category": category,
                "danger": danger,
                "photos": photos,
            }
        )

    runs: list[dict[str, Any]] = []
    measurement_id = 0
    for run_index in range(TOTAL_RUNS):
        cycle_count = (run_index + 1) * CYCLE_STEP
        test_time = (start_time + timedelta(hours=run_index * 2)).isoformat()
        measurements = []
        for point in points:
            measurement_id += 1
            point_seq = int(point["point_id"])
            category = point["category"]
            max_strain, min_strain = _strain_for(category, point_seq, run_index, TOTAL_RUNS)
            amplitude = round((max_strain - min_strain) / 2, 2)
            is_abnormal = amplitude > ABNORMAL_AMPLITUDE_UE
            reason = f"应变幅 {amplitude} με 超过阈值 {ABNORMAL_AMPLITUDE_UE:.0f} με" if is_abnormal else None
            measurements.append(
                {
                    "id": measurement_id,
                    "point_id": point["point_id"],
                    "max_strain_ue": max_strain,
                    "min_strain_ue": min_strain,
                    "mean_strain_ue": round((max_strain + min_strain) / 2, 2),
                    "amplitude_strain_ue": amplitude,
                    "range_strain_ue": round(max_strain - min_strain, 2),
                    "is_abnormal": is_abnormal,
                    "abnormal_reason": reason,
                    "remark": f"类别{category}-{CATEGORY_META[category]['label']}",
                }
            )
        runs.append(
            {
                "id": run_index + 1,
                "run_name": f"RUN-{run_index + 1:02d}",
                "cycle_count": cycle_count,
                "test_time": test_time,
                "remark": f"第 {run_index + 1} 轮台架疲劳加载",
                "measurements": measurements,
            }
        )

    # 人工标注异常：测点 05 在第 7 轮疑似裂纹扩展
    run7 = runs[6]
    for measurement in run7["measurements"]:
        if measurement["point_id"] == "05":
            measurement["is_abnormal"] = True
            measurement["abnormal_reason"] = "人工复核：应变幅突增，疑似裂纹扩展"

    # 裂缝记录：测点 05-08 在第 4/7/10 轮各一条
    crack_records: list[dict[str, Any]] = []
    crack_id = 0
    for run_index in (3, 6, 9):
        run = runs[run_index]
        for point in points:
            if point["point_id"] < "05":
                continue
            crack_id += 1
            cycle_count = run["cycle_count"]
            path = f"cracks/{point['point_id']}/{point['point_id']}_cycle_{cycle_count}_001.svg"
            package_files[path] = crack_svg(point["point_id"], cycle_count, crack_id)
            crack_records.append(
                {
                    "id": crack_id,
                    "point_id": point["point_id"],
                    "test_run_id": run["id"],
                    "run_cycle_count": cycle_count,
                    "cycle_count": cycle_count,
                    "path": path,
                    "filename": Path(path).name,
                    "original_filename": Path(path).name,
                    "content_type": "image/svg+xml",
                    "sha256": "",
                    "remark": f"裂缝观测：{CATEGORY_META[point['category']]['label']}点位",
                }
            )

    # Dewesoft 导入记录：第 1 轮与第 10 轮
    dewesoft_imports: list[dict[str, Any]] = []
    for dewe_seq, run_index in ((0, 0), (1, TOTAL_RUNS - 1)):
        run = runs[run_index]
        data_path = f"dewesoft/{run['run_name']}_frame.dat"
        channel_names = [point["point_id"] for point in points] + ["TEMP-01"]
        package_files[data_path] = (
            "$$ PointProcess 示例 Dewesoft 导出数据（占位内容）\n"
            f"$$ run: {run['run_name']}  cycle: {run['cycle_count']}\n"
            f"$$ channels: {','.join(channel_names)}\n"
            "$$ ...\n"
        ).encode("utf-8")
        measurement_by_point = {m["point_id"]: m for m in run["measurements"]}
        channels = []
        for point in points:
            measurement = measurement_by_point[point["point_id"]]
            channels.append(
                {
                    "id": len(channels) + 1,
                    "channel_name": point["point_id"],
                    "unit": "ue",
                    "sample_count": 60000,
                    "matched_point_id": point["point_id"],
                    "measurement_id": measurement["id"],
                    "stable_min_strain_ue": measurement["min_strain_ue"],
                    "stable_max_strain_ue": measurement["max_strain_ue"],
                    "stable_mean_strain_ue": measurement["mean_strain_ue"],
                    "raw_json": json.dumps({"device": "Dewesoft", "demo": True}, ensure_ascii=False),
                }
            )
        channels.append(
            {
                "id": len(channels) + 1,
                "channel_name": "TEMP-01",
                "unit": "degC",
                "sample_count": 60000,
                "matched_point_id": None,
                "measurement_id": None,
                "stable_min_strain_ue": None,
                "stable_max_strain_ue": None,
                "stable_mean_strain_ue": None,
                "raw_json": None,
            }
        )
        dewesoft_imports.append(
            {
                "id": dewe_seq + 1,
                "test_run_id": run["id"],
                "cycle_count": run["cycle_count"],
                "run_name": run["run_name"],
                "filename": Path(data_path).name,
                "path": data_path,
                "status": "imported",
                "message": None,
                "duration_seconds": 62.5,
                "stable_start_seconds": 10.0,
                "stable_end_seconds": 55.0,
                "matched_channel_count": len(points),
                "unmatched_channel_count": 1,
                "raw_metadata_json": json.dumps(
                    {"device": "Dewesoft", "sample_rate_hz": 1200, "demo": True}, ensure_ascii=False
                ),
                "channels": channels,
            }
        )

    return {
        "points": points,
        "runs": runs,
        "crack_records": crack_records,
        "dewesoft_imports": dewesoft_imports,
        "package_files": package_files,
    }


# ═══════════════════════════════════════════════════════════════
#  manifest.json / pointprocess_backup.json / records.xlsx
# ═══════════════════════════════════════════════════════════════

def build_manifest(dataset: dict[str, Any], now: str, fem_stats: dict[str, int]) -> dict[str, Any]:
    manifest_points = []
    for point in dataset["points"]:
        manifest_points.append(
            {
                "point_id": point["point_id"],
                "point_name": point["point_name"],
                "point_type": point["point_type"],
                "component": point["component"],
                "side": point["side"],
                "position_description": point["position_description"],
                "direction": point["direction"],
                "bridge_type": point["bridge_type"],
                "resistance_ohm": point["resistance_ohm"],
                "install_status": point["install_status"],
                "check_status": point["check_status"],
                "channel": {
                    "device": "Dewesoft",
                    "channel_name": point["point_id"],
                    "unit": "ue",
                    "sample_rate_hz": 1200.0,
                    "remark": f"Category {point['category']} — {CATEGORY_META[point['category']]['label']}",
                },
                "cae_mapping": {
                    "cae_point_id": f"CAE_{point['point_id']}",
                    "cae_component": point["component"],
                    "cae_result_type": "strain",
                    "danger_level": point["danger"],
                    "remark": CATEGORY_META[point["category"]]["desc"],
                },
                "photos": [
                    {
                        "photo_id": photo["photo_id"],
                        "type": photo["type"],
                        "path": photo["path"],
                        "filename": photo["filename"],
                        "taken_time": photo["taken_time"],
                        "sha256": photo["sha256"],
                        "remark": photo["remark"],
                    }
                    for photo in point["photos"]
                ],
                "tags": TAGS_MAP[point["category"]],
                "remark": f"类别{point['category']} — {CATEGORY_META[point['category']]['desc']}",
                "created_time": now,
                "updated_time": now,
                "custom_fields": {"behavior_category": point["category"]},
            }
        )
    return {
        "schema_version": "1.0.0",
        "export_info": {
            "export_id": EXPORT_ID,
            "export_time": now,
            "app_name": "PointProcess Web",
            "app_version": "1.0",
            "device_name": "Demo Generator",
            "operator": "Lee Chao",
            "remark": "PointProcess 全要素演示数据包（点位 + 测量数据 + 裂缝 + Dewesoft + FEM）",
        },
        "project": {
            "project_id": PROJECT_ID,
            "project_name": PROJECT_NAME,
            "test_object": "车架",
            "test_type": "疲劳试验",
            "department": "整车试验部",
            "vehicle_or_product": "非公路工程车辆车架",
            "test_stage": "全要素演示验证",
            "description": (
                "演示项目：包含 8 个点位（4 类行为模式）、10 轮测量数据、12 条裂缝记录、"
                "2 条 Dewesoft 导入记录以及带 INCLUDE 子文件与部件分组的 FEM 模型，"
                "可通过完整备份导入一次性恢复全部内容。"
            ),
            "created_time": now,
            "updated_time": now,
        },
        "points": manifest_points,
        "files": [
            {
                "file_id": "POINTPROCESS-RECORDS-XLSX",
                "type": "analysis_workbook",
                "path": "records.xlsx",
                "filename": "records.xlsx",
                "sha256": "",
                "remark": "演示记录工作簿",
            },
            {
                "file_id": "POINTPROCESS-FEM-MODEL",
                "type": "fem_model",
                "path": "fem",
                "filename": "fem",
                "sha256": "",
                "remark": "项目 FEM 模型源文件（导入时自动解析渲染）",
            },
        ],
        "custom_fields": {"pointprocess_backup": "pointprocess_backup.json"},
    }


def build_backup(
    dataset: dict[str, Any],
    manifest: dict[str, Any],
    now: str,
    fem_stats: dict[str, int],
    main_filename: str = "frame_assembly.fem",
) -> dict[str, Any]:
    manifest_point_by_id = {point["point_id"]: point for point in manifest["points"]}
    backup_points = []
    for point_seq, point in enumerate(dataset["points"], start=1):
        manifest_point = manifest_point_by_id[point["point_id"]]
        backup_points.append(
            {
                "id": point_seq,
                "point_id": point["point_id"],
                "point_name": point["point_name"],
                "point_type": point["point_type"],
                "component": point["component"],
                "side": point["side"],
                "position_description": point["position_description"],
                "direction": point["direction"],
                "bridge_type": point["bridge_type"],
                "resistance_ohm": point["resistance_ohm"],
                "install_status": point["install_status"],
                "check_status": point["check_status"],
                "remark": manifest_point["remark"],
                "raw_json": json.dumps(manifest_point, ensure_ascii=False),
                "created_at": now,
                "updated_at": now,
                "channel": manifest_point["channel"],
                "cae_mapping": manifest_point["cae_mapping"],
                "photos": [
                    {
                        "id": point_seq * 10 + photo_seq,
                        "photo_id": photo["photo_id"],
                        "type": photo["type"],
                        "path": photo["path"],
                        "original_path": photo["path"],
                        "filename": photo["filename"],
                        "original_filename": photo["filename"],
                        "taken_time": photo["taken_time"],
                        "sha256": "",
                        "remark": photo["remark"],
                    }
                    for photo_seq, photo in enumerate(point["photos"])
                ],
            }
        )

    return {
        "format": "pointprocess_project_backup",
        "version": "1.0",
        "export_id": EXPORT_ID,
        "exported_at": now,
        "project": {
            "id": 1,
            "project_id": PROJECT_ID,
            "project_name": PROJECT_NAME,
            "test_object": "车架",
            "test_type": "疲劳试验",
            "department": "整车试验部",
            "vehicle_or_product": "非公路工程车辆车架",
            "test_stage": "全要素演示验证",
            "description": manifest["project"]["description"],
            "source_export_id": None,
            "source_export_time": None,
            "raw_manifest_json": json.dumps(manifest, ensure_ascii=False),
            "created_at": now,
            "updated_at": now,
        },
        "points": backup_points,
        "test_runs": [
            {
                "id": run["id"],
                "run_name": run["run_name"],
                "cycle_count": run["cycle_count"],
                "test_time": run["test_time"],
                "remark": run["remark"],
                "created_at": now,
                "measurements": [
                    {
                        "id": measurement["id"],
                        "point_id": measurement["point_id"],
                        "max_strain_ue": measurement["max_strain_ue"],
                        "min_strain_ue": measurement["min_strain_ue"],
                        "mean_strain_ue": measurement["mean_strain_ue"],
                        "amplitude_strain_ue": measurement["amplitude_strain_ue"],
                        "range_strain_ue": measurement["range_strain_ue"],
                        "is_abnormal": measurement["is_abnormal"],
                        "abnormal_reason": measurement["abnormal_reason"],
                        "remark": measurement["remark"],
                        "created_at": now,
                        "updated_at": now,
                    }
                    for measurement in run["measurements"]
                ],
            }
            for run in dataset["runs"]
        ],
        "crack_records": dataset["crack_records"],
        "dewesoft_imports": dataset["dewesoft_imports"],
        "fem_model": {
            "main_filename": main_filename,
            "source_name": main_filename,
            "node_count": fem_stats["node_count"],
            "element_count": fem_stats["element_count"],
            "triangle_count": fem_stats["element_count"] * 2,
            "status": "ready",
            "error_message": None,
            "artifact_version": None,
        },
    }


def build_records_xlsx(dataset: dict[str, Any], manifest: dict[str, Any]) -> bytes:
    import io

    from openpyxl import Workbook

    wb = Workbook()
    summary = wb.active
    summary.title = "项目概览"
    summary.append(["字段", "值"])
    project = manifest["project"]
    for key, label in (
        ("project_id", "项目ID"),
        ("project_name", "项目名称"),
        ("test_object", "测试对象"),
        ("test_type", "试验类型"),
        ("department", "部门"),
        ("vehicle_or_product", "产品/车型"),
        ("test_stage", "试验阶段"),
    ):
        summary.append([label, project[key]])
    summary.append(["点位数量", len(dataset["points"])])
    summary.append(["测试轮次数量", len(dataset["runs"])])
    summary.append(["裂缝记录数量", len(dataset["crack_records"])])

    point_sheet = wb.create_sheet("点位清单")
    point_sheet.append(["点位编号", "点位名称", "类型", "部件", "方位", "位置描述", "方向", "桥路", "电阻", "安装状态", "检查状态", "备注"])
    for point in manifest["points"]:
        point_sheet.append(
            [
                point["point_id"], point["point_name"], point["point_type"], point["component"],
                point["side"], point["position_description"], point["direction"], point["bridge_type"],
                point["resistance_ohm"], point["install_status"], point["check_status"], point["remark"],
            ]
        )

    crack_sheet = wb.create_sheet("裂缝照片")
    crack_sheet.append(["点位编号", "循环次数", "轮次", "导出路径", "备注"])
    for crack in dataset["crack_records"]:
        crack_sheet.append([crack["point_id"], crack["cycle_count"], f"RUN-{crack['test_run_id']:02d}", crack["path"], crack["remark"]])

    stress_sheet = wb.create_sheet("应力变化情况")
    stress_sheet.append(["点位编号", "点位名称", "部件", "方向"] + [f"{run['cycle_count']}次应力幅(MPa)" for run in dataset["runs"]])
    modulus = 206000.0
    for point in manifest["points"]:
        values = []
        for run in dataset["runs"]:
            measurement = next(m for m in run["measurements"] if m["point_id"] == point["point_id"])
            values.append(round(measurement["amplitude_strain_ue"] * modulus / 1e6, 2))
        stress_sheet.append([point["point_id"], point["point_name"], point["component"], point["direction"], *values])

    for run in dataset["runs"]:
        sheet = wb.create_sheet(str(run["cycle_count"]))
        sheet.append(["点位编号", "点位名称", "最大应变(ue)", "最小应变(ue)", "应变幅(ue)", "应变范围(ue)", "应力幅(MPa)", "是否异常", "异常原因"])
        for measurement in run["measurements"]:
            point = next(p for p in manifest["points"] if p["point_id"] == measurement["point_id"])
            sheet.append(
                [
                    measurement["point_id"], point["point_name"],
                    measurement["max_strain_ue"], measurement["min_strain_ue"],
                    measurement["amplitude_strain_ue"], measurement["range_strain_ue"],
                    round(measurement["amplitude_strain_ue"] * modulus / 1e6, 2),
                    measurement["is_abnormal"], measurement["abnormal_reason"],
                ]
            )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ═══════════════════════════════════════════════════════════════
#  打包
# ═══════════════════════════════════════════════════════════════

FULL_ZIP_NAME = "POINTPROCESS_DEMO_FULL_20260905.zip"
FULL_SUMMARY_NAME = "POINTPROCESS_DEMO_FULL_20260905_summary.json"
FEM_REPLACEMENT_ZIP_NAME = "POINTPROCESS_DEMO_COMPLEX_FEM_REPLACEMENT_20260905.zip"
FEM_REPLACEMENT_SUMMARY_NAME = "POINTPROCESS_DEMO_COMPLEX_FEM_REPLACEMENT_20260905_summary.json"
SIMPLE_FEM_MAIN = "frame_assembly_simple.fem"
COMPLEX_FEM_MAIN = "frame_assembly_v2.fem"


def main() -> None:
    parser = argparse.ArgumentParser(description="生成 PointProcess 全要素演示数据包")
    parser.add_argument("--out", default=str(ROOT / "sample_data"), help="输出目录（默认 sample_data/）")
    parser.add_argument("--seed", type=int, default=20260905, help="随机种子")
    parser.add_argument(
        "--mode",
        choices=("full", "fem-replacement"),
        default="full",
        help="full=全要素演示包（内置第一版骨架 FEM，供先导入项目后整体替换演示）；"
        "fem-replacement=单独生成第二版高保真复杂 FEM 替换包",
    )
    args = parser.parse_args()

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    if args.mode == "fem-replacement":
        _build_fem_replacement_package(out_dir)
        return
    _build_full_package(out_dir, args.seed)


def _stage_and_validate(out_dir: Path, fem_files: dict[str, str], main_filename: str) -> dict[str, int]:
    """把 FEM 源文件写入临时目录，用后端解析器实际解析并返回统计。"""
    import shutil

    staging = out_dir / ".fem_validate_tmp"
    shutil.rmtree(staging, ignore_errors=True)
    for relative, text in fem_files.items():
        target = staging / relative
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(text, encoding="utf-8")
    try:
        return validate_fem_source(staging, main_filename=main_filename)
    finally:
        shutil.rmtree(staging, ignore_errors=True)


def _build_full_package(out_dir: Path, seed: int) -> None:
    """完整备份导入包：项目 + 点位/测量/裂缝/Dewesoft + 第一版骨架 FEM。"""
    random.seed(seed)
    tz = timezone(timedelta(hours=8))
    now_dt = datetime(2026, 9, 5, 10, 0, tzinfo=tz)
    now = now_dt.isoformat()
    start_time = now_dt

    dataset = build_dataset(now, start_time)

    # FEM 源文件：第一版骨架模型（先导入项目，演示后续用复杂模型整体替换）
    fem_files = build_simple_fem_source_files()
    fem_stats = _stage_and_validate(out_dir, fem_files, SIMPLE_FEM_MAIN)
    print(f"FEM 校验通过: {fem_stats}")
    for relative, text in fem_files.items():
        dataset["package_files"][f"fem/source/{relative}"] = text.encode("utf-8")

    manifest = build_manifest(dataset, now, fem_stats)
    backup = build_backup(dataset, manifest, now, fem_stats, main_filename=SIMPLE_FEM_MAIN)
    dataset["package_files"]["records.xlsx"] = build_records_xlsx(dataset, manifest)
    dataset["package_files"]["raw/readme.txt"] = "示例原始数据目录（演示数据）\n".encode("utf-8")
    dataset["package_files"]["attachments/试验大纲.txt"] = (
        "车架疲劳台架试验大纲（演示数据）\n\n"
        "1. 加载波形：正弦波，频率 5 Hz\n"
        "2. 目标循环次数：80,000 次\n"
        "3. 数据采集：Dewesoft，采样率 1200 Hz\n"
        "4. 每 8,000 次循环记录一轮应变极值并巡检裂缝\n"
    ).encode("utf-8")

    zip_path = out_dir / FULL_ZIP_NAME
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as package:
        package.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
        package.writestr("pointprocess_backup.json", json.dumps(backup, ensure_ascii=False, indent=2))
        for relative, content in sorted(dataset["package_files"].items()):
            package.writestr(relative, content)

    summary = {
        "zip_path": str(zip_path),
        "project_id": PROJECT_ID,
        "fem_mode": "simple (第一版骨架模型, frame_assembly_simple.fem)",
        "point_count": len(dataset["points"]),
        "photo_count": sum(len(point["photos"]) for point in dataset["points"]),
        "test_run_count": len(dataset["runs"]),
        "measurement_count": sum(len(run["measurements"]) for run in dataset["runs"]),
        "abnormal_measurement_count": sum(
            1 for run in dataset["runs"] for m in run["measurements"] if m["is_abnormal"]
        ),
        "crack_record_count": len(dataset["crack_records"]),
        "dewesoft_import_count": len(dataset["dewesoft_imports"]),
        "dewesoft_channel_count": sum(len(item["channels"]) for item in dataset["dewesoft_imports"]),
        "fem_stats": fem_stats,
        "package_file_count": 2 + len(dataset["package_files"]),
        "zip_size_bytes": zip_path.stat().st_size,
        "coverage": [
            "project create/list/detail via full backup import",
            "8 points with channels, CAE mappings, photos and tags",
            "10 fatigue runs and 80 measurement records (4 behavior patterns)",
            "manual and threshold-based abnormal measurement cases",
            "12 crack timeline records linked to points and runs",
            "2 Dewesoft import records with matched/unmatched channels",
            "raw/ and attachments/ folder copy",
            "FEM model (simple skeleton) with INCLUDE sub-file and component grouping",
            "replace demo: import this package, then upload the complex FEM replacement package",
        ],
        "import_guide": "系统内选择 导入 → 上传该 zip → 预览 → 确认导入，即可一次性恢复全部内容（含 FEM 模型）",
        "replacement_guide": "在项目 FEM 预览页点击「重新导入 .fem 文件」，选择复杂替换包解压后的 frame_assembly_v2.fem，即可替换为复杂模型并重新渲染",
    }
    summary_path = out_dir / FULL_SUMMARY_NAME
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"ZIP     → {zip_path}")
    print(f"SUMMARY → {summary_path}")
    print(
        f"点位 {summary['point_count']} | 照片 {summary['photo_count']} | 轮次 {summary['test_run_count']} "
        f"| 测量 {summary['measurement_count']}（异常 {summary['abnormal_measurement_count']}）"
        f"| 裂缝 {summary['crack_record_count']} | Dewesoft {summary['dewesoft_import_count']}"
    )


def _build_fem_replacement_package(out_dir: Path) -> None:
    """单独生成第二版（复杂）FEM 单文件替换包。

    包内是自包含的 ``frame_assembly_v2.fem``（纵梁等全部内联，无 INCLUDE），
    可在 FEM 预览页直接选择上传：把项目内第一版骨架模型替换为复杂模型并重新渲染。
    """
    import shutil
    import tempfile

    fem_files = build_complex_fem_source_files(single_file=True)
    assert list(fem_files) == [COMPLEX_FEM_MAIN]
    with tempfile.TemporaryDirectory() as td:
        folder = Path(td)
        for relative, text in fem_files.items():
            target = folder / relative
            target.parent.mkdir(parents=True, exist_ok=True)
            target.write_text(text, encoding="utf-8")
        fem_stats = validate_fem_source(folder, main_filename=COMPLEX_FEM_MAIN, include_count=0)
    print(f"复杂 FEM（单文件）校验通过: {fem_stats}")

    readme = (
        "PointProcess 复杂 FEM 替换文件（第二版高保真车架模型）\n"
        "====================================================\n\n"
        "用途：整体替换项目内已导入/已渲染的 FEM 模型。\n\n"
        "操作步骤：\n"
        "  1) 先导入演示全要素包 POINTPROCESS_DEMO_FULL_20260905.zip（内含第一版骨架模型）；\n"
        "  2) 打开该项目的模型预览页；\n"
        "  3) 点击「重新导入 .fem 文件」，选择本目录中的 frame_assembly_v2.fem 上传；\n"
        "  4) 后端解析渲染完成后，页面模型即被替换为复杂模型并重新渲染。\n\n"
        "模型内容（与第一版对比）：\n"
        "  - 单一自包含文件（无 INCLUDE 依赖，可直接单独上传）\n"
        "  - 箱型截面左右纵梁（上/下盖板 + 内/外腹板，多 PID 板厚）\n"
        "  - 8 根等距横梁（PID 交替 5.0/6.0）、前后保险杠、四角立柱与顶部封口\n"
        "  - CROD 连杆（圆形截面 PBARL 属性）、CTRIA3 过渡封口\n"
        "  - 钢材 + 铝合金两种材料，约 {node} 节点 / {elem} 单元\n"
        "".format(node=fem_stats["node_count"], elem=fem_stats["element_count"])
    ).encode("utf-8")

    zip_path = out_dir / FEM_REPLACEMENT_ZIP_NAME
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as package:
        for relative, text in fem_files.items():
            package.writestr(relative, text)
        package.writestr("README-复杂FEM替换说明.txt", readme)

    summary = {
        "zip_path": str(zip_path),
        "fem_mode": "complex single-file (第二版高保真模型, frame_assembly_v2.fem)",
        "fem_stats": fem_stats,
        "file_count": 2 + len(fem_files),
        "zip_size_bytes": zip_path.stat().st_size,
        "usage_guide": "解压后选择 frame_assembly_v2.fem 单个文件，在 FEM 预览页「重新导入 .fem 文件」上传",
    }
    summary_path = out_dir / FEM_REPLACEMENT_SUMMARY_NAME
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"ZIP     → {zip_path}")
    print(f"SUMMARY → {summary_path}")


if __name__ == "__main__":
    main()
