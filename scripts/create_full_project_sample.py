from __future__ import annotations

import json
import math
import shutil
import sys
import tempfile
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT / "backend"
OUT_DIR = ROOT / "sample_data"
ZIP_PATH = OUT_DIR / "POINTPROCESS_FULL_PROJECT_STRESS_20260628.zip"
SUMMARY_PATH = OUT_DIR / "POINTPROCESS_FULL_PROJECT_STRESS_20260628_summary.json"

POINT_COUNT = 72
RUN_COUNT = 36
DEWESOFT_IMPORT_COUNT = 6
CRACK_RECORD_COUNT = 120
EXPORT_TIME = datetime(2026, 6, 28, 14, 0, tzinfo=timezone(timedelta(hours=8)))


def iso(offset_minutes: int = 0) -> str:
    return (EXPORT_TIME + timedelta(minutes=offset_minutes)).isoformat()


def svg_image(title: str, subtitle: str, fill: str, accent: str) -> bytes:
    return f"""<svg xmlns="http://www.w3.org/2000/svg" width="960" height="640" viewBox="0 0 960 640">
<rect width="960" height="640" fill="{fill}"/>
<rect x="54" y="54" width="852" height="532" fill="#ffffff" opacity="0.92"/>
<path d="M120 430 C230 280 330 430 430 250 S650 220 790 390" fill="none" stroke="{accent}" stroke-width="18" stroke-linecap="round"/>
<circle cx="210" cy="360" r="24" fill="{accent}"/>
<circle cx="518" cy="252" r="24" fill="{accent}"/>
<circle cx="742" cy="366" r="24" fill="{accent}"/>
<text x="480" y="190" font-family="Arial" font-size="58" text-anchor="middle" fill="#172026">{title}</text>
<text x="480" y="252" font-family="Arial" font-size="28" text-anchor="middle" fill="#52646c">{subtitle}</text>
<text x="480" y="540" font-family="Arial" font-size="24" text-anchor="middle" fill="#6b7d85">PointProcess synthetic import test asset</text>
</svg>""".encode("utf-8")


def sha256_hex(data: bytes) -> str:
    import hashlib

    return hashlib.sha256(data).hexdigest()


def measurement_values(point_index: int, run_index: int) -> tuple[float, float, bool, str | None]:
    base = 70 + (point_index % 18) * 8.5
    growth = 1 + run_index * (0.015 + (point_index % 9) * 0.0025)
    wave = 1 + math.sin((point_index + run_index) / 4) * 0.06
    if point_index in {9, 18, 27, 36, 45, 54, 63, 72} and run_index > 20:
        growth *= 1.22 + (run_index - 20) * 0.018
    amplitude = base * growth * wave
    mean = (point_index % 11) * 2.4 + run_index * 0.7
    max_strain = round(mean + amplitude, 3)
    min_strain = round(mean - amplitude, 3)
    abnormal = amplitude > 360 or (point_index in {12, 24, 48, 60} and run_index in {11, 12, 13, 14})
    reason = None
    if abnormal:
        reason = "fast amplitude growth" if amplitude > 360 else "manual review marker"
    return max_strain, min_strain, abnormal, reason


def stress_fields(max_strain: float | None, min_strain: float | None) -> dict[str, float | None]:
    if max_strain is None or min_strain is None:
        return {
            "mean_strain_ue": None,
            "amplitude_strain_ue": None,
            "range_strain_ue": None,
            "stress_max_mpa": None,
            "stress_min_mpa": None,
            "stress_mean_mpa": None,
            "stress_amplitude_mpa": None,
            "stress_range_mpa": None,
        }
    mean = (max_strain + min_strain) / 2
    amplitude = (max_strain - min_strain) / 2
    value_range = max_strain - min_strain
    return {
        "mean_strain_ue": round(mean, 6),
        "amplitude_strain_ue": round(amplitude, 6),
        "range_strain_ue": round(value_range, 6),
        "stress_max_mpa": round(max_strain * 0.206, 6),
        "stress_min_mpa": round(min_strain * 0.206, 6),
        "stress_mean_mpa": round(mean * 0.206, 6),
        "stress_amplitude_mpa": round(amplitude * 0.206, 6),
        "stress_range_mpa": round(value_range * 0.206, 6),
    }


def build_package() -> tuple[dict[str, Any], dict[str, bytes], dict[str, Any]]:
    components = [
        "front_crossmember",
        "rear_crossmember",
        "left_longitudinal_beam",
        "right_longitudinal_beam",
        "battery_tray",
        "suspension_mount",
        "welded_joint",
        "reinforcement_plate",
    ]
    sides = ["left", "right", "front", "rear", "middle", "upper", "lower", "unknown"]
    directions = ["longitudinal", "transverse", "vertical", "principal", "rosette_0", "rosette_45", "rosette_90", "unknown"]
    bridge_types = ["1/4_bridge", "1/2_bridge", "full_bridge", "unknown"]
    install_statuses = ["installed", "planned", "damaged", "removed", "abandoned"]
    check_statuses = ["checked", "unchecked", "abnormal", "rechecked"]
    danger_levels = ["low", "medium", "high", "critical", "unknown"]
    photo_types = ["overall", "local", "wiring"]

    files: dict[str, bytes] = {}
    points: list[dict[str, Any]] = []
    manifest_points: list[dict[str, Any]] = []

    for index in range(1, POINT_COUNT + 1):
        point_code = f"{index:02d}"
        component = components[(index - 1) % len(components)]
        photos: list[dict[str, Any]] = []
        for photo_index, photo_type in enumerate(photo_types, start=1):
            path = f"photos/{point_code}/{point_code}_{photo_type}_{photo_index:02d}.svg"
            data = svg_image(point_code, photo_type, "#edf3f4" if photo_type == "overall" else "#f4efe1", "#2f7a8a" if photo_type != "wiring" else "#b55d2d")
            files[path] = data
            photos.append(
                {
                    "id": index * 10 + photo_index,
                    "photo_id": f"PHOTO-{point_code}-{photo_index:02d}",
                    "type": photo_type,
                    "path": path,
                    "original_path": f"field/{point_code}/{photo_type}.jpg",
                    "filename": Path(path).name,
                    "original_filename": Path(path).name,
                    "taken_time": iso(index + photo_index),
                    "sha256": sha256_hex(data),
                    "remark": f"{photo_type} photo for import and media preview tests",
                }
            )

        point = {
            "id": index,
            "point_id": point_code,
            "point_name": f"Frame fatigue point {point_code}",
            "point_type": "strain_gauge" if index % 10 else "displacement_sensor",
            "component": component,
            "side": sides[(index - 1) % len(sides)],
            "position_description": f"{component} test location {index}, includes routing and weld context.",
            "direction": directions[(index - 1) % len(directions)],
            "bridge_type": bridge_types[(index - 1) % len(bridge_types)],
            "resistance_ohm": round(119.6 + index * 0.035, 3),
            "install_status": install_statuses[(index - 1) % len(install_statuses)],
            "check_status": check_statuses[(index - 1) % len(check_statuses)],
            "remark": f"Coverage point {index}: edit, filter, risk and media test data.",
            "raw_json": json.dumps(
                {
                    "tags": ["fatigue", component, "risk" if index % 9 == 0 else "normal"],
                    "custom_fields": {"fixture": f"fixture-{(index - 1) % 6 + 1}", "weld_distance_mm": 12 + index},
                    "created_time": iso(-120 + index),
                    "updated_time": iso(index),
                },
                ensure_ascii=False,
            ),
            "created_at": iso(-120 + index),
            "updated_at": iso(index),
            "channel": {
                "device": "Dewesoft SIRIUS",
                "channel_name": point_code,
                "unit": "ue",
                "sample_rate_hz": 1000 + (index % 4) * 500,
                "remark": f"channel mapped to {point_code}",
            },
            "cae_mapping": {
                "cae_point_id": f"CAE-{point_code}",
                "cae_component": f"CAE_{component}",
                "cae_result_type": "strain",
                "danger_level": danger_levels[(index - 1) % len(danger_levels)],
                "remark": "synthetic CAE mapping",
            },
            "photos": photos,
        }
        points.append(point)
        manifest_points.append(
            {
                "point_id": point["point_id"],
                "point_name": point["point_name"],
                "point_type": point["point_type"],
                "component": point["component"],
                "side": point["side"],
                "position_description": point["position_description"],
                "direction": point["direction"],
                "bridge_type": point["bridge_type"],
                "resistance_ohm": point["resistance_ohm"],
                "install_status": point["install_status"],
                "check_status": point["check_status"],
                "channel": point["channel"],
                "cae_mapping": point["cae_mapping"],
                "photos": [
                    {
                        "photo_id": photo["photo_id"],
                        "type": photo["type"],
                        "path": photo["path"],
                        "filename": photo["filename"],
                        "taken_time": photo["taken_time"],
                        "sha256": photo["sha256"],
                        "remark": photo["remark"],
                    }
                    for photo in photos
                ],
                "tags": json.loads(point["raw_json"])["tags"],
                "remark": point["remark"],
                "created_time": point["created_at"],
                "updated_time": point["updated_at"],
                "custom_fields": json.loads(point["raw_json"])["custom_fields"],
            }
        )

    test_runs: list[dict[str, Any]] = []
    measurement_old_id = 1
    measurement_id_by_run_point: dict[tuple[int, str], int] = {}
    abnormal_points: set[str] = set()
    for run_index in range(1, RUN_COUNT + 1):
        cycle_count = run_index * 5000
        measurements: list[dict[str, Any]] = []
        for point_index, point in enumerate(points, start=1):
            max_strain, min_strain, abnormal, reason = measurement_values(point_index, run_index)
            fields = stress_fields(max_strain, min_strain)
            if abnormal:
                abnormal_points.add(point["point_id"])
            measurement = {
                "id": measurement_old_id,
                "point_id": point["point_id"],
                "max_strain_ue": max_strain,
                "min_strain_ue": min_strain,
                **fields,
                "is_abnormal": abnormal,
                "abnormal_reason": reason,
                "remark": f"run {run_index} point {point['point_id']} synthetic measurement",
                "created_at": iso(run_index * 8),
                "updated_at": iso(run_index * 8 + 2),
            }
            measurements.append(measurement)
            measurement_id_by_run_point[(run_index, point["point_id"])] = measurement_old_id
            measurement_old_id += 1
        test_runs.append(
            {
                "id": run_index,
                "run_name": f"R{run_index:02d}-{cycle_count}cycles",
                "cycle_count": cycle_count,
                "test_time": iso(run_index * 60),
                "remark": "generated fatigue run with complete point coverage",
                "created_at": iso(run_index * 60),
                "measurements": measurements,
            }
        )

    crack_records: list[dict[str, Any]] = []
    crack_candidates = [point for point in points if int(point["point_id"]) % 3 == 0]
    for crack_index in range(1, CRACK_RECORD_COUNT + 1):
        point = crack_candidates[(crack_index - 1) % len(crack_candidates)]
        run_index = 8 + ((crack_index * 5) % (RUN_COUNT - 8))
        cycle_count = test_runs[run_index - 1]["cycle_count"]
        path = f"cracks/{point['point_id']}/{point['point_id']}_cycle_{cycle_count}_{crack_index:03d}.svg"
        data = svg_image(f"CRACK {crack_index:03d}", f"{point['point_id']} cycle {cycle_count}", "#f8ece9", "#9d2f2f")
        files[path] = data
        crack_records.append(
            {
                "id": crack_index,
                "point_id": point["point_id"],
                "test_run_id": run_index,
                "run_cycle_count": cycle_count,
                "cycle_count": cycle_count,
                "path": path,
                "filename": Path(path).name,
                "original_filename": Path(path).name,
                "content_type": "image/svg+xml",
                "sha256": sha256_hex(data),
                "remark": "crack timeline marker for overview chart and crack page tests",
                "created_at": iso(run_index * 60 + crack_index),
                "updated_at": iso(run_index * 60 + crack_index + 1),
            }
        )

    dewesoft_imports: list[dict[str, Any]] = []
    for import_index in range(1, DEWESOFT_IMPORT_COUNT + 1):
        run_index = import_index * 5
        run = test_runs[run_index - 1]
        path = f"dewesoft/import_{import_index:02d}_cycle_{run['cycle_count']}.csv"
        lines = ["time_s,channel,value_ue"]
        for sample_index in range(1, 121):
            point = points[(sample_index + import_index) % len(points)]
            value = round(20 + sample_index * 0.9 + import_index * 4 + math.sin(sample_index / 8) * 12, 4)
            lines.append(f"{sample_index / 100:.2f},{point['point_id']},{value}")
        data = ("\n".join(lines) + "\n").encode("utf-8")
        files[path] = data

        channels: list[dict[str, Any]] = []
        for point_index, point in enumerate(points, start=1):
            matched = point_index % 17 != 0
            channel_name = point["point_id"] if matched else f"UNMATCHED-{import_index}-{point_index}"
            channels.append(
                {
                    "id": import_index * 1000 + point_index,
                    "channel_name": channel_name,
                    "unit": "ue",
                    "sample_count": 1200 + point_index,
                    "matched_point_id": point["point_id"] if matched else None,
                    "measurement_id": measurement_id_by_run_point.get((run_index, point["point_id"])) if matched else None,
                    "stable_min_strain_ue": round(-80 - point_index * 0.7, 3),
                    "stable_max_strain_ue": round(110 + point_index * 1.1, 3),
                    "stable_mean_strain_ue": round(12 + point_index * 0.18, 3),
                    "raw_json": json.dumps({"quality": "ok" if matched else "unmatched", "source": path}),
                    "created_at": iso(import_index * 75 + point_index),
                }
            )
        dewesoft_imports.append(
            {
                "id": import_index,
                "test_run_id": run_index,
                "cycle_count": run["cycle_count"],
                "run_name": run["run_name"],
                "filename": Path(path).name,
                "path": path,
                "status": "imported" if import_index != DEWESOFT_IMPORT_COUNT else "warning",
                "message": "synthetic dewesoft import with matched and unmatched channels",
                "duration_seconds": 180.5 + import_index,
                "stable_start_seconds": 20.0,
                "stable_end_seconds": 140.0,
                "matched_channel_count": sum(1 for channel in channels if channel["matched_point_id"]),
                "unmatched_channel_count": sum(1 for channel in channels if not channel["matched_point_id"]),
                "raw_metadata_json": json.dumps({"sample_rate_hz": 1000, "source": "synthetic"}),
                "created_at": iso(import_index * 75),
                "channels": channels,
            }
        )

    workbook_bytes = build_workbook(points, test_runs, crack_records, dewesoft_imports)
    files["records.xlsx"] = workbook_bytes
    files["raw/load_profile.csv"] = build_load_profile().encode("utf-8")
    files["raw/chamber_temperature.csv"] = build_temperature_profile().encode("utf-8")
    files["attachments/test_plan.txt"] = (
        "PointProcess full project sample test plan\n"
        "- import preview and confirm\n"
        "- project list/detail/overview\n"
        "- point media overall/local/wiring\n"
        "- measurement trend and risk summary\n"
        "- crack records and chart markers\n"
        "- Dewesoft import history and channel matching\n"
        "- JSON/CSV/full ZIP export\n"
    ).encode("utf-8")
    files["attachments/operator_notes.txt"] = (
        "Synthetic data set. Values are generated for broad UI and API coverage, not physical validation.\n"
    ).encode("utf-8")

    project = {
        "id": 1,
        "project_id": "POINTPROCESS-FULL-STRESS-20260628",
        "project_name": "PointProcess full function stress sample",
        "test_object": "vehicle frame assembly",
        "test_type": "fatigue durability bench test",
        "department": "validation lab",
        "vehicle_or_product": "EV chassis platform",
        "test_stage": "system integration regression",
        "description": "Large import package covering points, media, measurements, abnormal analysis, cracks, Dewesoft records, attachments and re-export.",
        "source_export_id": "WEB-FULL-STRESS-20260628",
        "source_export_time": EXPORT_TIME.isoformat(),
        "raw_manifest_json": None,
        "created_at": iso(-180),
        "updated_at": iso(240),
    }
    manifest = {
        "schema_version": "1.0.0",
        "export_info": {
            "export_id": "WEB-FULL-STRESS-20260628",
            "export_time": EXPORT_TIME.isoformat(),
            "app_name": "PointProcess Web",
            "app_version": "1.0",
            "operator": "Codex synthetic data generator",
            "remark": "Full function import package with pointprocess backup.",
        },
        "project": {
            "project_id": project["project_id"],
            "project_name": project["project_name"],
            "test_object": project["test_object"],
            "test_type": project["test_type"],
            "department": project["department"],
            "vehicle_or_product": project["vehicle_or_product"],
            "test_stage": project["test_stage"],
            "description": project["description"],
            "created_time": project["created_at"],
            "updated_time": project["updated_at"],
        },
        "points": manifest_points,
        "files": [
            {
                "file_id": "POINTPROCESS-RECORDS-XLSX",
                "type": "analysis_workbook",
                "path": "records.xlsx",
                "filename": "records.xlsx",
                "sha256": sha256_hex(files["records.xlsx"]),
                "remark": "Workbook with project summary and generated measurement data.",
            }
        ],
        "custom_fields": {"pointprocess_backup": "pointprocess_backup.json"},
    }
    backup = {
        "format": "pointprocess_project_backup",
        "version": "1.0",
        "export_id": "WEB-FULL-STRESS-20260628",
        "exported_at": EXPORT_TIME.isoformat(),
        "project": project,
        "points": points,
        "test_runs": test_runs,
        "crack_records": crack_records,
        "dewesoft_imports": dewesoft_imports,
    }
    project["raw_manifest_json"] = json.dumps(manifest, ensure_ascii=False)

    summary = {
        "zip_path": str(ZIP_PATH),
        "project_id": project["project_id"],
        "point_count": len(points),
        "photo_count": sum(len(point["photos"]) for point in points),
        "test_run_count": len(test_runs),
        "measurement_count": sum(len(run["measurements"]) for run in test_runs),
        "abnormal_point_count": len(abnormal_points),
        "crack_record_count": len(crack_records),
        "dewesoft_import_count": len(dewesoft_imports),
        "dewesoft_channel_count": sum(len(item["channels"]) for item in dewesoft_imports),
        "file_count": len(files) + 2,
        "coverage": [
            "project create/list/detail/update/delete data paths",
            "manifest import preview",
            "full backup import confirm",
            "point filters, statuses, CAE mappings, channels and tags",
            "overall/local/wiring point media",
            "36 fatigue runs and 2592 measurement records",
            "manual and computed abnormal measurement cases",
            "analysis summary, max amplitude and growth ranking",
            "crack timeline records and crack image retrieval",
            "Dewesoft import history, matched channels and unmatched channels",
            "raw files and attachments folder copy",
            "full project ZIP re-export",
        ],
    }
    return {"manifest": manifest, "backup": backup}, files, summary


def build_workbook(
    points: list[dict[str, Any]],
    test_runs: list[dict[str, Any]],
    crack_records: list[dict[str, Any]],
    dewesoft_imports: list[dict[str, Any]],
) -> bytes:
    wb = Workbook()
    summary = wb.active
    summary.title = "summary"
    summary.append(["metric", "value"])
    summary.append(["points", len(points)])
    summary.append(["runs", len(test_runs)])
    summary.append(["measurements", sum(len(run["measurements"]) for run in test_runs)])
    summary.append(["cracks", len(crack_records)])
    summary.append(["dewesoft_imports", len(dewesoft_imports)])

    point_sheet = wb.create_sheet("points")
    point_sheet.append(["point_id", "point_name", "component", "side", "direction", "bridge_type", "install_status", "check_status"])
    for point in points:
        point_sheet.append([point["point_id"], point["point_name"], point["component"], point["side"], point["direction"], point["bridge_type"], point["install_status"], point["check_status"]])

    measurement_sheet = wb.create_sheet("measurements_sample")
    measurement_sheet.append(["run_name", "cycle_count", "point_id", "max_strain_ue", "min_strain_ue", "amplitude_strain_ue", "is_abnormal", "abnormal_reason"])
    for run in test_runs[:6]:
        for record in run["measurements"][:24]:
            measurement_sheet.append([run["run_name"], run["cycle_count"], record["point_id"], record["max_strain_ue"], record["min_strain_ue"], record["amplitude_strain_ue"], record["is_abnormal"], record["abnormal_reason"]])

    crack_sheet = wb.create_sheet("cracks")
    crack_sheet.append(["point_id", "cycle_count", "path", "remark"])
    for record in crack_records:
        crack_sheet.append([record["point_id"], record["cycle_count"], record["path"], record["remark"]])

    dewesoft_sheet = wb.create_sheet("dewesoft")
    dewesoft_sheet.append(["filename", "cycle_count", "matched", "unmatched", "status"])
    for item in dewesoft_imports:
        dewesoft_sheet.append([item["filename"], item["cycle_count"], item["matched_channel_count"], item["unmatched_channel_count"], item["status"]])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)
    try:
        wb.save(tmp_path)
        return tmp_path.read_bytes()
    finally:
        tmp_path.unlink(missing_ok=True)


def build_load_profile() -> str:
    lines = ["time_s,load_kn,actuator_mm"]
    for index in range(1, 721):
        lines.append(f"{index * 0.5:.1f},{45 + math.sin(index / 16) * 8:.4f},{2.5 + math.cos(index / 11) * 0.7:.4f}")
    return "\n".join(lines) + "\n"


def build_temperature_profile() -> str:
    lines = ["time_s,chamber_c,frame_c"]
    for index in range(1, 721):
        lines.append(f"{index * 0.5:.1f},{25 + math.sin(index / 40) * 3:.3f},{27 + math.sin(index / 35) * 2.2:.3f}")
    return "\n".join(lines) + "\n"


def write_zip(payloads: dict[str, Any], files: dict[str, bytes]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(ZIP_PATH, "w", zipfile.ZIP_DEFLATED) as package:
        package.writestr("manifest.json", json.dumps(payloads["manifest"], ensure_ascii=False, indent=2))
        package.writestr("pointprocess_backup.json", json.dumps(payloads["backup"], ensure_ascii=False, indent=2))
        for path in sorted(files):
            package.writestr(path, files[path])


def validate_zip(summary: dict[str, Any]) -> dict[str, Any]:
    sys.path.insert(0, str(BACKEND_DIR))
    from sqlalchemy import create_engine, select
    from sqlalchemy.orm import sessionmaker

    from app import models
    from app.database import Base
    from app.services import import_service
    from app.services.analysis_service import summary_for_project
    from app.services import project_export_service
    from app.utils.zip_utils import safe_extract, validate_zip_members

    temp_root = Path(tempfile.mkdtemp(prefix="pointprocess-full-sample-"))
    try:
        storage_dir = temp_root / "storage"
        for folder in ["imports", "projects", "dewesoft", "temp"]:
            (storage_dir / folder).mkdir(parents=True, exist_ok=True)
        import_service.PROJECT_ROOT = temp_root
        import_service.STORAGE_DIR = storage_dir
        project_export_service.PROJECT_ROOT = temp_root
        project_export_service.STORAGE_DIR = storage_dir

        db_path = temp_root / "sample.db"
        engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False}, future=True)
        Base.metadata.create_all(bind=engine)
        SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)

        with zipfile.ZipFile(ZIP_PATH) as package:
            member_errors = validate_zip_members(package)
            if member_errors:
                raise RuntimeError("; ".join(member_errors))
            extract_dir = storage_dir / "temp" / "TMP-FULLSAMPLE" / "extract"
            safe_extract(package, extract_dir)

        temp_dir = extract_dir.parent
        zip_copy = temp_dir / ZIP_PATH.name
        shutil.copy2(ZIP_PATH, zip_copy)
        with SessionLocal() as db:
            preview = import_service._build_preview_from_extract(db, "TMP-FULLSAMPLE", temp_dir, extract_dir, ZIP_PATH.name, zip_copy)
            if not preview.can_import:
                raise RuntimeError(f"preview failed: {preview.errors}")
            project = import_service.confirm_import(db, "TMP-FULLSAMPLE")

            actual_counts = {
                "points": db.query(models.TestPoint).filter(models.TestPoint.project_db_id == project.id).count(),
                "media": db.query(models.MediaFile).filter(models.MediaFile.project_db_id == project.id).count(),
                "runs": db.query(models.TestRun).filter(models.TestRun.project_db_id == project.id).count(),
                "measurements": db.query(models.MeasurementRecord).join(models.TestRun).filter(models.TestRun.project_db_id == project.id).count(),
                "cracks": db.query(models.CrackRecord).filter(models.CrackRecord.project_db_id == project.id).count(),
                "dewesoft_imports": db.query(models.DewesoftImport).filter(models.DewesoftImport.project_db_id == project.id).count(),
                "dewesoft_channels": db.query(models.DewesoftChannel).join(models.DewesoftImport).filter(models.DewesoftImport.project_db_id == project.id).count(),
            }
            expected_counts = {
                "points": summary["point_count"],
                "media": summary["photo_count"],
                "runs": summary["test_run_count"],
                "measurements": summary["measurement_count"],
                "cracks": summary["crack_record_count"],
                "dewesoft_imports": summary["dewesoft_import_count"],
                "dewesoft_channels": summary["dewesoft_channel_count"],
            }
            if actual_counts != expected_counts:
                raise RuntimeError(f"count mismatch: expected={expected_counts}, actual={actual_counts}")

            analysis = summary_for_project(db, project.id)
            if analysis["measurement_count"] != summary["measurement_count"]:
                raise RuntimeError("analysis summary did not see all measurements")
            if analysis["abnormal_count"] <= 0:
                raise RuntimeError("analysis summary did not include abnormal points")

            export_path, _ = project_export_service.build_project_export_zip(db, project.id)
            with zipfile.ZipFile(export_path) as exported:
                exported_names = set(exported.namelist())
                if "pointprocess_backup.json" not in exported_names or "records.xlsx" not in exported_names:
                    raise RuntimeError("round-trip export is missing full backup files")
                exported_backup = json.loads(exported.read("pointprocess_backup.json").decode("utf-8"))
            round_trip_counts = {
                "points": len(exported_backup["points"]),
                "runs": len(exported_backup["test_runs"]),
                "cracks": len(exported_backup["crack_records"]),
                "dewesoft_imports": len(exported_backup["dewesoft_imports"]),
            }

        return {
            "validated": True,
            "preview": {
                "can_import": preview.can_import,
                "point_count": preview.point_count,
                "photo_count": preview.photo_count,
                "warnings": preview.warnings,
                "errors": preview.errors,
            },
            "import_counts": actual_counts,
            "analysis": {
                "point_count": analysis["point_count"],
                "run_count": analysis["run_count"],
                "measurement_count": analysis["measurement_count"],
                "abnormal_count": analysis["abnormal_count"],
                "max_amplitude_rows": len(analysis["max_amplitude_points"]),
                "fastest_growth_rows": len(analysis["fastest_growth_points"]),
            },
            "round_trip_export_counts": round_trip_counts,
        }
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def main() -> None:
    payloads, files, summary = build_package()
    write_zip(payloads, files)
    validation = validate_zip(summary)
    summary["zip_size_bytes"] = ZIP_PATH.stat().st_size
    summary["validation"] = validation
    SUMMARY_PATH.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(ZIP_PATH)
    print(SUMMARY_PATH)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
