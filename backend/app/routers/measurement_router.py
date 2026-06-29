import json
import uuid
from collections import Counter, defaultdict
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app import models
from app.database import STORAGE_DIR, get_db
from app.schemas import (
    MeasurementBatchCreate,
    MeasurementCreate,
    MeasurementOut,
    MeasurementUpdate,
    PointMeasurementRowCreate,
    PointMeasurementRowOut,
    PointMeasurementRowsSave,
    PointMeasurementRowSave,
    PointMeasurementRowUpdate,
    TestRunOut,
    TestRunUpdate,
    XlsxImportConfirmRequest,
    XlsxImportPreview,
    XlsxImportResult,
    XlsxImportRowError,
    XlsxImportStrategyEnum,
    XlsxPreviewItem,
    XlsxRowStatus,
)
from app.services.analysis_service import compute_measurement_fields, refresh_point_abnormal_flags


router = APIRouter(tags=["measurements"])

REQUIRED_XLSX_HEADERS = ["cycle_count", "point_id", "max_strain_ue", "min_strain_ue"]
"""XLSX 必要表头：cycle_count + point_id 唯一标识测量记录，max/min strain 为测量数据。

run_name 不再是必要表头 —— 缺失时自动生成 "{cycle_count}次"。
"""
OPTIONAL_XLSX_HEADERS = ["run_name", "test_time", "point_name", "remark", "data_source", "operator"]
"""XLSX 可选表头，提供更完整的导入数据。"""


def _validate_strain_consistency(max_strain: float | None, min_strain: float | None) -> None:
    """校验最大应变不小于最小应变。"""
    if max_strain is not None and min_strain is not None:
        if max_strain < min_strain:
            raise HTTPException(status_code=400, detail="最大应变不能小于最小应变")


def apply_measurement_payload(record: models.MeasurementRecord, payload: MeasurementCreate | MeasurementUpdate) -> None:
    data = payload.model_dump(exclude_unset=True)
    for field, value in data.items():
        setattr(record, field, value)
    # 校验数值一致性
    _validate_strain_consistency(record.max_strain_ue, record.min_strain_ue)
    compute_measurement_fields(record)
    if data.get("is_abnormal") is True:
        record.is_abnormal = True
        record.abnormal_reason = data.get("abnormal_reason") or "人工标记异常"


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_number(value: Any, row_number: int, field: str) -> float | None:
    text = _cell_text(value)
    if text == "":
        return None
    try:
        return float(text)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"第 {row_number} 行 {field} 不是数字: {text}") from exc


def _parse_xlsx_rows(file_bytes: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"无法读取 XLSX 文件: {exc}") from exc
    if not workbook.worksheets:
        raise HTTPException(status_code=400, detail="XLSX 文件中没有工作表")

    sheet = workbook["measurements"] if "measurements" in workbook.sheetnames else workbook.worksheets[0]
    headers = [_cell_text(cell.value) for cell in sheet[1]]
    header_index = {header: index for index, header in enumerate(headers) if header}
    missing = [header for header in REQUIRED_XLSX_HEADERS if header not in header_index]
    if missing:
        raise HTTPException(status_code=400, detail=f"模板缺少表头: {', '.join(missing)}")

    rows: list[dict[str, Any]] = []
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = {header: cells[index] if index < len(cells) else None for header, index in header_index.items()}
        if not any(_cell_text(row.get(field)) for field in ["max_strain_ue", "min_strain_ue", "remark"]):
            continue
        run_name = _cell_text(row.get("run_name"))
        cycle_count_text = _cell_text(row.get("cycle_count"))
        point_id = _cell_text(row.get("point_id"))
        if not run_name:
            raise HTTPException(status_code=400, detail=f"第 {row_number} 行缺少 run_name")
        if not cycle_count_text:
            raise HTTPException(status_code=400, detail=f"第 {row_number} 行缺少 cycle_count")
        try:
            cycle_count = int(float(cycle_count_text))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=f"第 {row_number} 行 cycle_count 不是数字: {cycle_count_text}") from exc
        if not point_id:
            raise HTTPException(status_code=400, detail=f"第 {row_number} 行缺少 point_id")

        rows.append(
            {
                "row_number": row_number,
                "run_name": run_name,
                "cycle_count": cycle_count,
                "test_time": _cell_text(row.get("test_time")) or None,
                "point_id": point_id,
                "max_strain_ue": _cell_number(row.get("max_strain_ue"), row_number, "max_strain_ue"),
                "min_strain_ue": _cell_number(row.get("min_strain_ue"), row_number, "min_strain_ue"),
                "remark": _cell_text(row.get("remark")) or None,
            }
        )
    if not rows:
        raise HTTPException(status_code=400, detail="没有可导入的数据，请至少填写最大应变或最小应变")
    return rows


def _create_or_update_measurement(
    db: Session,
    run_id: int,
    point_db_id: int,
    max_strain_ue: float | None,
    min_strain_ue: float | None,
    remark: str | None,
) -> models.MeasurementRecord:
    record = db.scalar(
        select(models.MeasurementRecord).where(
            models.MeasurementRecord.run_id == run_id,
            models.MeasurementRecord.point_db_id == point_db_id,
        )
    )
    if not record:
        record = models.MeasurementRecord(run_id=run_id, point_db_id=point_db_id)
    record.max_strain_ue = max_strain_ue
    record.min_strain_ue = min_strain_ue
    record.remark = remark
    _validate_strain_consistency(record.max_strain_ue, record.min_strain_ue)
    compute_measurement_fields(record)
    db.add(record)
    return record


def _measurement_row_out(record: models.MeasurementRecord) -> PointMeasurementRowOut:
    data = MeasurementOut.model_validate(record).model_dump()
    data["run_name"] = record.run.run_name
    data["cycle_count"] = record.run.cycle_count
    return PointMeasurementRowOut.model_validate(data)


def _measurement_payload_from_row(payload: PointMeasurementRowCreate | PointMeasurementRowUpdate | PointMeasurementRowSave) -> MeasurementUpdate:
    data = payload.model_dump(exclude_unset=True)
    return MeasurementUpdate(
        **{
            key: data[key]
            for key in ["max_strain_ue", "min_strain_ue", "is_abnormal", "abnormal_reason", "remark"]
            if key in data
        }
    )


def _upsert_point_measurement_row(
    db: Session,
    point: models.TestPoint,
    payload: PointMeasurementRowCreate | PointMeasurementRowUpdate | PointMeasurementRowSave,
    measurement_id: int | None = None,
) -> models.MeasurementRecord:
    data = payload.model_dump(exclude_unset=True)
    if measurement_id is None:
        if data.get("cycle_count") is None:
            raise HTTPException(status_code=400, detail="循环次数不能为空")
        cycle_count = data["cycle_count"]
        run = _find_run_by_cycle(db, point.project_db_id, cycle_count)
        if not run:
            run = models.TestRun(
                project_db_id=point.project_db_id,
                run_name=(data.get("run_name") or f"R{cycle_count}").strip() or f"R{cycle_count}",
                cycle_count=cycle_count,
            )
            db.add(run)
            db.flush()
        elif data.get("run_name") and _run_measurement_count(db, run.id) == 0:
            run.run_name = data["run_name"].strip() or run.run_name

        record = db.scalar(
            select(models.MeasurementRecord).where(
                models.MeasurementRecord.run_id == run.id,
                models.MeasurementRecord.point_db_id == point.id,
            )
        )
        if not record:
            record = models.MeasurementRecord(run_id=run.id, point_db_id=point.id)
        apply_measurement_payload(record, _measurement_payload_from_row(payload))
        db.add(record)
        db.flush()
        return record

    record = db.get(models.MeasurementRecord, measurement_id)
    if not record or record.point_db_id != point.id:
        raise HTTPException(status_code=404, detail="测量记录不存在")

    next_cycle_count = data["cycle_count"] if data.get("cycle_count") is not None else record.run.cycle_count
    next_run_name = data["run_name"].strip() if data.get("run_name") is not None else record.run.run_name
    next_run_name = next_run_name or record.run.run_name or f"R{next_cycle_count}"
    run_changed = next_cycle_count != record.run.cycle_count or next_run_name != record.run.run_name

    if run_changed:
        if next_cycle_count == record.run.cycle_count and _run_measurement_count(db, record.run_id, measurement_id) == 0:
            record.run.run_name = next_run_name
        else:
            target_run = _find_named_run(db, point.project_db_id, next_cycle_count, next_run_name)
            if not target_run and "run_name" not in data:
                target_run = _find_run_by_cycle(db, point.project_db_id, next_cycle_count)
            if not target_run:
                target_run = models.TestRun(
                    project_db_id=point.project_db_id,
                    run_name=next_run_name,
                    cycle_count=next_cycle_count,
                )
                db.add(target_run)
                db.flush()

            existing = db.scalar(
                select(models.MeasurementRecord).where(
                    models.MeasurementRecord.run_id == target_run.id,
                    models.MeasurementRecord.point_db_id == record.point_db_id,
                    models.MeasurementRecord.id != record.id,
                )
            )
            if existing:
                raise HTTPException(status_code=400, detail="目标循环次数下已存在该点位的测量记录，请先编辑或删除已有记录")
            record.run = target_run
    apply_measurement_payload(record, _measurement_payload_from_row(payload))
    db.flush()
    return record


def _run_measurement_count(db: Session, run_id: int, exclude_measurement_id: int | None = None) -> int:
    query = select(func.count()).select_from(models.MeasurementRecord).where(models.MeasurementRecord.run_id == run_id)
    if exclude_measurement_id is not None:
        query = query.where(models.MeasurementRecord.id != exclude_measurement_id)
    return db.scalar(query) or 0


def _find_named_run(db: Session, project_id: int, cycle_count: int, run_name: str) -> models.TestRun | None:
    return db.scalar(
        select(models.TestRun)
        .where(
            models.TestRun.project_db_id == project_id,
            models.TestRun.cycle_count == cycle_count,
            models.TestRun.run_name == run_name,
        )
        .order_by(models.TestRun.id)
    )


def _find_run_by_cycle(db: Session, project_id: int, cycle_count: int) -> models.TestRun | None:
    return db.scalar(
        select(models.TestRun)
        .where(models.TestRun.project_db_id == project_id, models.TestRun.cycle_count == cycle_count)
        .order_by(models.TestRun.id)
    )


@router.get("/api/test-runs/{run_id}", response_model=TestRunOut)
def get_test_run(run_id: int, db: Session = Depends(get_db)) -> TestRunOut:
    run = db.get(models.TestRun, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="测试轮次不存在")
    return TestRunOut.model_validate(run)


@router.put("/api/test-runs/{run_id}", response_model=TestRunOut)
def update_test_run(run_id: int, payload: TestRunUpdate, db: Session = Depends(get_db)) -> TestRunOut:
    run = db.get(models.TestRun, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="测试轮次不存在")
    data = payload.model_dump(exclude_unset=True)
    if "run_name" in data:
        run_name = (data["run_name"] or "").strip()
        if not run_name:
            raise HTTPException(status_code=400, detail="测试轮次名称不能为空")
        run.run_name = run_name
    if "cycle_count" in data and data["cycle_count"] is not None:
        run.cycle_count = data["cycle_count"]
    if "test_time" in data:
        run.test_time = data["test_time"]
    if "remark" in data:
        run.remark = data["remark"]
    db.commit()
    db.refresh(run)
    return TestRunOut.model_validate(run)


@router.delete("/api/test-runs/{run_id}")
def delete_test_run(run_id: int, db: Session = Depends(get_db)) -> dict:
    run = db.get(models.TestRun, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="测试轮次不存在")
    point_ids = [record.point_db_id for record in run.measurements]
    db.delete(run)
    db.flush()
    for point_id in set(point_ids):
        refresh_point_abnormal_flags(db, point_id)
    db.commit()
    return {"ok": True, "action": "permanently_deleted"}


@router.post("/api/test-runs/{run_id}/measurements", response_model=list[MeasurementOut])
def create_measurements(run_id: int, payload: MeasurementBatchCreate, db: Session = Depends(get_db)) -> list[MeasurementOut]:
    run = db.get(models.TestRun, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="测试轮次不存在")
    created: list[models.MeasurementRecord] = []
    for item in payload.measurements:
        point = db.get(models.TestPoint, item.point_db_id)
        if not point or point.project_db_id != run.project_db_id:
            raise HTTPException(status_code=400, detail=f"点位不属于当前项目: {item.point_db_id}")
        existing = db.scalar(
            select(models.MeasurementRecord).where(
                models.MeasurementRecord.run_id == run_id,
                models.MeasurementRecord.point_db_id == item.point_db_id,
            )
        )
        record = existing or models.MeasurementRecord(run_id=run_id, point_db_id=item.point_db_id)
        apply_measurement_payload(record, item)
        db.add(record)
        created.append(record)
    db.flush()
    for point_id in {record.point_db_id for record in created}:
        refresh_point_abnormal_flags(db, point_id)
    db.commit()
    for record in created:
        db.refresh(record)
    return [MeasurementOut.model_validate(record) for record in created]


@router.post("/api/projects/{project_id}/measurements/import-xlsx")
async def import_project_measurements_xlsx(
    project_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
) -> dict:
    """（已废弃）直接导入 XLSX 测量数据，上传即写入数据库。

    @deprecated 新前端请使用 preview → confirm 两步流程：
    - POST .../import-xlsx/preview  上传并预览
    - POST .../import-xlsx/confirm  确认后写入

    本接口保留以兼容旧版调用和脚本集成。
    """
    project = db.get(models.Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="请上传 XLSX/XLSM 文件，旧版 .xls 格式暂不支持")

    rows = _parse_xlsx_rows(await file.read())
    points = db.execute(select(models.TestPoint).where(models.TestPoint.project_db_id == project_id)).scalars().all()
    point_by_id = {point.point_id: point for point in points}
    for row in rows:
        if row["point_id"] not in point_by_id:
            raise HTTPException(status_code=400, detail=f"第 {row['row_number']} 行点位编号不存在: {row['point_id']}")

    groups: dict[tuple[str, int, str | None], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        groups[(row["run_name"], row["cycle_count"], row["test_time"])].append(row)

    created_runs = 0
    measurement_count = 0
    affected_point_ids: set[int] = set()
    for (run_name, cycle_count, test_time), group_rows in sorted(groups.items(), key=lambda item: item[0][1]):
        run = db.scalar(
            select(models.TestRun).where(
                models.TestRun.project_db_id == project_id,
                models.TestRun.run_name == run_name,
                models.TestRun.cycle_count == cycle_count,
            )
        )
        if not run:
            run = models.TestRun(
                project_db_id=project_id,
                run_name=run_name,
                cycle_count=cycle_count,
                test_time=test_time,
                remark=f"XLSX import: {file.filename}",
            )
            db.add(run)
            db.flush()
            created_runs += 1
        for row in group_rows:
            point = point_by_id[row["point_id"]]
            record = _create_or_update_measurement(
                db,
                run.id,
                point.id,
                row["max_strain_ue"],
                row["min_strain_ue"],
                row["remark"],
            )
            measurement_count += 1
            affected_point_ids.add(record.point_db_id)

    db.flush()
    for point_id in affected_point_ids:
        refresh_point_abnormal_flags(db, point_id)
    db.commit()
    return {
        "ok": True,
        "run_count": len(groups),
        "created_run_count": created_runs,
        "measurement_count": measurement_count,
    }


@router.get("/api/test-runs/{run_id}/measurements", response_model=list[MeasurementOut])
def list_run_measurements(run_id: int, db: Session = Depends(get_db)) -> list[MeasurementOut]:
    records = db.execute(
        select(models.MeasurementRecord).where(models.MeasurementRecord.run_id == run_id).order_by(models.MeasurementRecord.point_db_id)
    ).scalars()
    return [MeasurementOut.model_validate(record) for record in records]


@router.get("/api/points/{point_id}/measurements", response_model=list[MeasurementOut])
def list_point_measurements(point_id: int, db: Session = Depends(get_db)) -> list[MeasurementOut]:
    records = db.execute(
        select(models.MeasurementRecord)
        .join(models.TestRun)
        .where(models.MeasurementRecord.point_db_id == point_id)
        .order_by(models.TestRun.cycle_count, models.TestRun.id)
    ).scalars()
    return [MeasurementOut.model_validate(record) for record in records]


@router.get("/api/points/{point_id}/measurement-rows", response_model=list[PointMeasurementRowOut])
def list_point_measurement_rows(point_id: int, db: Session = Depends(get_db)) -> list[PointMeasurementRowOut]:
    records = db.execute(
        select(models.MeasurementRecord)
        .join(models.TestRun)
        .where(models.MeasurementRecord.point_db_id == point_id)
        .order_by(models.TestRun.cycle_count, models.TestRun.id)
    ).scalars()
    return [_measurement_row_out(record) for record in records]


@router.post("/api/points/{point_id}/measurement-rows", response_model=PointMeasurementRowOut)
def create_point_measurement_row(
    point_id: int,
    payload: PointMeasurementRowCreate,
    db: Session = Depends(get_db),
) -> PointMeasurementRowOut:
    point = db.get(models.TestPoint, point_id)
    if not point:
        raise HTTPException(status_code=404, detail="点位不存在")
    record = _upsert_point_measurement_row(db, point, payload)
    db.flush()
    refresh_point_abnormal_flags(db, point.id)
    db.commit()
    db.refresh(record)
    return _measurement_row_out(record)


@router.put("/api/points/{point_id}/measurement-rows/{measurement_id}", response_model=PointMeasurementRowOut)
def update_point_measurement_row(
    point_id: int,
    measurement_id: int,
    payload: PointMeasurementRowUpdate,
    db: Session = Depends(get_db),
) -> PointMeasurementRowOut:
    record = db.get(models.MeasurementRecord, measurement_id)
    if not record or record.point_db_id != point_id:
        raise HTTPException(status_code=404, detail="测量记录不存在")
    record = _upsert_point_measurement_row(db, record.point, payload, measurement_id)
    db.flush()
    refresh_point_abnormal_flags(db, record.point_db_id)
    db.commit()
    db.refresh(record)
    return _measurement_row_out(record)


@router.put("/api/points/{point_id}/measurement-rows", response_model=list[PointMeasurementRowOut])
def save_point_measurement_rows(
    point_id: int,
    payload: PointMeasurementRowsSave,
    db: Session = Depends(get_db),
) -> list[PointMeasurementRowOut]:
    point = db.get(models.TestPoint, point_id)
    if not point:
        raise HTTPException(status_code=404, detail="点位不存在")
    deleted_ids = set(payload.deleted_measurement_ids)
    for measurement_id in deleted_ids:
        record = db.get(models.MeasurementRecord, measurement_id)
        if not record or record.point_db_id != point_id:
            raise HTTPException(status_code=404, detail=f"测量记录不存在: {measurement_id}")
        db.delete(record)
    db.flush()

    saved: list[models.MeasurementRecord] = []
    for row in payload.measurements:
        if row.id in deleted_ids:
            raise HTTPException(status_code=400, detail=f"测量记录不能同时删除和保存: {row.id}")
        saved.append(_upsert_point_measurement_row(db, point, row, row.id))

    db.flush()
    refresh_point_abnormal_flags(db, point_id)
    db.commit()
    for record in saved:
        db.refresh(record)
    return [_measurement_row_out(record) for record in saved]


@router.delete("/api/points/{point_id}/measurement-rows/{measurement_id}")
def delete_point_measurement_row(point_id: int, measurement_id: int, db: Session = Depends(get_db)) -> dict:
    record = db.get(models.MeasurementRecord, measurement_id)
    if not record or record.point_db_id != point_id:
        raise HTTPException(status_code=404, detail="测量记录不存在")
    db.delete(record)
    db.flush()
    refresh_point_abnormal_flags(db, point_id)
    db.commit()
    return {"ok": True, "action": "permanently_deleted"}


@router.put("/api/measurements/{measurement_id}", response_model=MeasurementOut)
def update_measurement(measurement_id: int, payload: MeasurementUpdate, db: Session = Depends(get_db)) -> MeasurementOut:
    record = db.get(models.MeasurementRecord, measurement_id)
    if not record:
        raise HTTPException(status_code=404, detail="测量记录不存在")
    apply_measurement_payload(record, payload)
    db.flush()
    refresh_point_abnormal_flags(db, record.point_db_id)
    db.commit()
    db.refresh(record)
    return MeasurementOut.model_validate(record)


@router.delete("/api/measurements/{measurement_id}")
def delete_measurement(measurement_id: int, db: Session = Depends(get_db)) -> dict:
    record = db.get(models.MeasurementRecord, measurement_id)
    if not record:
        raise HTTPException(status_code=404, detail="测量记录不存在")
    point_id = record.point_db_id
    db.delete(record)
    db.flush()
    refresh_point_abnormal_flags(db, point_id)
    db.commit()
    return {"ok": True, "action": "permanently_deleted"}


# ── XLSX 导入：预检 + 确认 两步流程 ────────────────────────────────────────

XLSX_TEMP_DIR = STORAGE_DIR / "temp" / "xlsx"


def _parse_xlsx_rows_lenient(file_bytes: bytes) -> tuple[list[dict[str, Any]], list[XlsxImportRowError]]:
    """解析 XLSX 行，收集所有错误而非遇错即停。

    身份规则：
    - 测量记录唯一身份: cycle_count + point_id
    - TestRun 唯一身份: project_db_id + cycle_count
    - point_name 仅用于人工核对，不参与点位匹配

    有效行规则：
    - cycle_count 非空且为 >= 0 的整数
    - point_id 非空
    - max_strain_ue 或 min_strain_ue 至少一个非空
    - 仅填写 max 或 min 其中一个 → warning
    - max_strain_ue >= min_strain_ue
    """
    errors: list[XlsxImportRowError] = []
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"无法读取 XLSX 文件: {exc}") from exc
    if not workbook.worksheets:
        raise HTTPException(status_code=400, detail="XLSX 文件中没有工作表")

    sheet = workbook["measurements"] if "measurements" in workbook.sheetnames else workbook.worksheets[0]
    headers = [_cell_text(cell.value) for cell in sheet[1]]
    header_index = {header: index for index, header in enumerate(headers) if header}
    missing = [header for header in REQUIRED_XLSX_HEADERS if header not in header_index]
    if missing:
        raise HTTPException(status_code=400, detail=f"模板缺少必要表头: {', '.join(missing)}。\n"
                                                    f"必要表头: {', '.join(REQUIRED_XLSX_HEADERS)}。\n"
                                                    f"可选表头: {', '.join(OPTIONAL_XLSX_HEADERS)}。")

    rows: list[dict[str, Any]] = []
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = {header: cells[index] if index < len(cells) else None for header, index in header_index.items()}

        # 尝试获取所有已知字段
        cycle_count_text = _cell_text(row.get("cycle_count"))
        point_id = _cell_text(row.get("point_id"))
        max_strain_text = _cell_text(row.get("max_strain_ue"))
        min_strain_text = _cell_text(row.get("min_strain_ue"))
        remark_text = _cell_text(row.get("remark"))

        # 跳过完全空行（所有关键字段都为空）
        if not any([cycle_count_text, point_id, max_strain_text, min_strain_text, remark_text]):
            continue

        # 逐字段校验
        row_errors: list[XlsxImportRowError] = []

        # cycle_count 校验
        cycle_count: int | None = None
        if not cycle_count_text:
            row_errors.append(XlsxImportRowError(row=row_number, field="cycle_count", message="缺少循环次数", severity="error"))
        else:
            try:
                cycle_count = int(float(cycle_count_text))
                if cycle_count < 0:
                    row_errors.append(XlsxImportRowError(row=row_number, field="cycle_count", message="循环次数不能为负数", severity="error"))
            except (ValueError, OverflowError):
                row_errors.append(XlsxImportRowError(row=row_number, field="cycle_count", message=f"循环次数不是有效整数: {cycle_count_text}", severity="error"))

        # point_id 校验
        if not point_id:
            row_errors.append(XlsxImportRowError(row=row_number, field="point_id", message="缺少点位编号", severity="error"))

        # max/min strain 校验
        max_strain: float | None = None
        min_strain: float | None = None
        has_max = bool(max_strain_text)
        has_min = bool(min_strain_text)

        if not has_max and not has_min:
            row_errors.append(XlsxImportRowError(row=row_number, field="max_strain_ue", message="最大应变和最小应变至少需要填写一个", severity="error"))
        else:
            if has_max:
                try:
                    max_strain = float(max_strain_text)
                except ValueError:
                    row_errors.append(XlsxImportRowError(row=row_number, field="max_strain_ue", message=f"最大应变不是有效数字: {max_strain_text}", severity="error"))
            if has_min:
                try:
                    min_strain = float(min_strain_text)
                except ValueError:
                    row_errors.append(XlsxImportRowError(row=row_number, field="min_strain_ue", message=f"最小应变不是有效数字: {min_strain_text}", severity="error"))

            # 如果两个都成功解析，检查一致性
            if max_strain is not None and min_strain is not None:
                if max_strain < min_strain:
                    row_errors.append(XlsxImportRowError(row=row_number, field="max_strain_ue", message="最大应变不能小于最小应变", severity="error"))

            # 仅填写了其中一个 → warning
            if has_max != has_min:
                missing_field = "min_strain_ue" if has_max else "max_strain_ue"
                row_errors.append(XlsxImportRowError(row=row_number, field=missing_field, message="仅填写了最大或最小应变之一，数据不完整", severity="warning"))

        errors.extend(row_errors)

        run_name = _cell_text(row.get("run_name"))
        test_time = _cell_text(row.get("test_time")) or None
        point_name = _cell_text(row.get("point_name")) or None
        data_source = _cell_text(row.get("data_source")) or None
        operator = _cell_text(row.get("operator")) or None

        rows.append({
            "row_number": row_number,
            "run_name": run_name or None,
            "cycle_count": cycle_count if cycle_count is not None else 0,
            "test_time": test_time,
            "point_id": point_id or None,
            "point_name": point_name,
            "max_strain_ue": max_strain,
            "min_strain_ue": min_strain,
            "remark": remark_text or None,
            "data_source": data_source,
            "operator": operator,
            "has_errors": len(row_errors) > 0 and any(e.severity == "error" for e in row_errors),
        })

    if not rows:
        raise HTTPException(status_code=400, detail="没有可导入的数据，请至少填写循环次数、点位编号和应变值（最大或最小）")
    return rows, errors


def _build_xlsx_preview(db: Session, project_id: int, rows: list[dict], errors: list[XlsxImportRowError], filename: str = "") -> XlsxImportPreview:
    """根据解析结果构建详细预检报告。

    身份规则：
    - TestRun 唯一身份: project_db_id + cycle_count
    - MeasurementRecord 唯一身份: run_id + point_db_id (等价于 project_id + cycle_count + point_id)
    - 点位匹配仅使用 point_id，point_name 不参与匹配

    行状态分类：
    - INVALID: 有 error 级别的校验错误
    - UNKNOWN_POINT: point_id 在当前项目中不存在
    - FILE_DUPLICATE: 同一文件内 (cycle_count, point_id) 出现多次
    - EXISTING_MEASUREMENT: 数据库中已存在对应的 MeasurementRecord
    - NEW_MEASUREMENT: 数据库中不存在，可新增
    """
    # 加载当前项目点位
    points = db.execute(select(models.TestPoint).where(models.TestPoint.project_db_id == project_id)).scalars().all()
    point_by_id: dict[str, models.TestPoint] = {point.point_id: point for point in points}

    # 加载当前项目已有 TestRun（按 cycle_count 索引）
    existing_runs = db.execute(
        select(models.TestRun).where(models.TestRun.project_db_id == project_id)
    ).scalars().all()
    run_by_cycle: dict[int, models.TestRun] = {run.cycle_count: run for run in existing_runs}

    # 分离无效行（有 error 级别错误）
    error_row_numbers: set[int] = {e.row for e in errors if e.severity == "error"}
    warning_errors = [e for e in errors if e.severity == "warning"]

    # 检测文件内重复: (cycle_count, point_id) 组合
    cycle_point_pairs: list[tuple[int, str]] = []
    for row in rows:
        if row["row_number"] not in error_row_numbers and row["point_id"] and row.get("cycle_count", 0) > 0:
            cycle_point_pairs.append((row["cycle_count"], row["point_id"]))
    cycle_point_counter = Counter(cycle_point_pairs)
    duplicate_pairs: set[tuple[int, str]] = {pair for pair, count in cycle_point_counter.items() if count > 1}

    # 构建预览项
    items: list[XlsxPreviewItem] = []
    seen_pairs: set[tuple[int, str]] = set()
    warnings_from_preview: list[XlsxImportRowError] = []

    # 统计涉及的 cycle_count（有效行）
    xlsx_cycle_counts: set[int] = set()
    new_run_cycles: set[int] = set()
    existing_run_cycles: set[int] = set()

    for row in rows:
        row_num = row["row_number"]
        point_id = row.get("point_id") or ""
        cycle_count = row.get("cycle_count", 0)

        # 无效行
        if row_num in error_row_numbers:
            row_errs = [e for e in errors if e.row == row_num]
            items.append(XlsxPreviewItem(
                row_index=row_num,
                cycle_count=cycle_count if cycle_count > 0 else None,
                point_id=point_id or None,
                point_name=row.get("point_name"),
                run_name=row.get("run_name"),
                test_time=row.get("test_time"),
                max_strain_ue=row.get("max_strain_ue"),
                min_strain_ue=row.get("min_strain_ue"),
                status=XlsxRowStatus.INVALID,
                message="; ".join(e.message for e in row_errs),
            ))
            continue

        # 未知点位
        point = point_by_id.get(point_id)
        if not point:
            items.append(XlsxPreviewItem(
                row_index=row_num,
                cycle_count=cycle_count,
                point_id=point_id,
                point_name=row.get("point_name"),
                run_name=row.get("run_name"),
                test_time=row.get("test_time"),
                max_strain_ue=row.get("max_strain_ue"),
                min_strain_ue=row.get("min_strain_ue"),
                status=XlsxRowStatus.UNKNOWN_POINT,
                message=f"点位 {point_id} 不存在于当前项目",
            ))
            continue

        # 文件内重复（第二个及之后出现的标记为重复）
        pair = (cycle_count, point_id)
        if pair in duplicate_pairs:
            if pair in seen_pairs:
                items.append(XlsxPreviewItem(
                    row_index=row_num,
                    cycle_count=cycle_count,
                    point_id=point_id,
                    point_name=point.point_name,
                    run_name=row.get("run_name"),
                    test_time=row.get("test_time"),
                    max_strain_ue=row.get("max_strain_ue"),
                    min_strain_ue=row.get("min_strain_ue"),
                    status=XlsxRowStatus.FILE_DUPLICATE,
                    message=f"文件内重复: cycle_count={cycle_count}, point_id={point_id}",
                ))
                continue
            seen_pairs.add(pair)

        # 记录首次出现的 pair
        if pair not in duplicate_pairs:
            pass  # 不是重复
        else:
            seen_pairs.add(pair)

        # 统计 cycle_count
        xlsx_cycle_counts.add(cycle_count)
        if cycle_count in run_by_cycle:
            existing_run_cycles.add(cycle_count)
        else:
            new_run_cycles.add(cycle_count)

        # point_name 不一致检测
        if row.get("point_name") and row["point_name"] != point.point_name:
            warnings_from_preview.append(XlsxImportRowError(
                row=row_num,
                field="point_name",
                message=f"XLSX 中点位名称 '{row['point_name']}' 与系统中 '{point.point_name}' 不一致，将以 point_id 匹配为准",
                severity="warning",
            ))

        # run_name 不一致检测（已有轮次）
        if cycle_count in run_by_cycle:
            existing_run = run_by_cycle[cycle_count]
            if row.get("run_name") and row["run_name"] != existing_run.run_name:
                warnings_from_preview.append(XlsxImportRowError(
                    row=row_num,
                    field="run_name",
                    message=f"XLSX 中轮次名称 '{row['run_name']}' 与已有轮次 '{existing_run.run_name}' 不一致，默认保留已有轮次名称",
                    severity="warning",
                ))
            if row.get("test_time") and row["test_time"] != (existing_run.test_time or ""):
                warnings_from_preview.append(XlsxImportRowError(
                    row=row_num,
                    field="test_time",
                    message=f"XLSX 中测试时间与已有轮次不一致，默认保留已有时间",
                    severity="warning",
                ))

        # 检测数据库中是否已有记录
        existing_record: models.MeasurementRecord | None = None
        if cycle_count in run_by_cycle:
            run = run_by_cycle[cycle_count]
            existing_record = db.scalar(
                select(models.MeasurementRecord).where(
                    models.MeasurementRecord.run_id == run.id,
                    models.MeasurementRecord.point_db_id == point.id,
                )
            )

        if existing_record:
            items.append(XlsxPreviewItem(
                row_index=row_num,
                cycle_count=cycle_count,
                point_id=point_id,
                point_name=point.point_name,
                run_name=row.get("run_name"),
                test_time=row.get("test_time"),
                max_strain_ue=row.get("max_strain_ue"),
                min_strain_ue=row.get("min_strain_ue"),
                status=XlsxRowStatus.EXISTING_MEASUREMENT,
                message="数据库中已有该记录",
                existing_max_strain_ue=existing_record.max_strain_ue,
                existing_min_strain_ue=existing_record.min_strain_ue,
                existing_run_name=existing_record.run.run_name if existing_record.run else None,
                incoming_max_strain_ue=row.get("max_strain_ue"),
                incoming_min_strain_ue=row.get("min_strain_ue"),
            ))
        else:
            items.append(XlsxPreviewItem(
                row_index=row_num,
                cycle_count=cycle_count,
                point_id=point_id,
                point_name=point.point_name,
                run_name=row.get("run_name"),
                test_time=row.get("test_time"),
                max_strain_ue=row.get("max_strain_ue"),
                min_strain_ue=row.get("min_strain_ue"),
                status=XlsxRowStatus.NEW_MEASUREMENT,
                message="新测量记录",
            ))

    # 汇总统计
    status_counter = Counter(item.status for item in items)
    total_rows = len(rows)
    valid_rows = status_counter.get(XlsxRowStatus.NEW_MEASUREMENT, 0) + status_counter.get(XlsxRowStatus.EXISTING_MEASUREMENT, 0)
    invalid_rows = status_counter.get(XlsxRowStatus.INVALID, 0)

    new_measurement_count = status_counter.get(XlsxRowStatus.NEW_MEASUREMENT, 0)
    existing_measurement_count = status_counter.get(XlsxRowStatus.EXISTING_MEASUREMENT, 0)
    unknown_point_count = status_counter.get(XlsxRowStatus.UNKNOWN_POINT, 0)
    file_duplicate_count = status_counter.get(XlsxRowStatus.FILE_DUPLICATE, 0)

    all_warnings = warning_errors + warnings_from_preview

    # can_confirm: 无 error 且无未知点位和文件内重复（默认安全策略）
    has_errors = len([e for e in errors if e.severity == "error"]) > 0
    can_confirm = not has_errors and unknown_point_count == 0 and file_duplicate_count == 0

    return XlsxImportPreview(
        preview_id="",  # 由调用方填充
        filename=filename,
        total_rows=total_rows,
        valid_rows=valid_rows,
        invalid_rows=invalid_rows,
        cycle_counts=sorted(xlsx_cycle_counts),
        new_run_count=len(new_run_cycles),
        existing_run_count=len(existing_run_cycles),
        new_measurement_count=new_measurement_count,
        existing_measurement_count=existing_measurement_count,
        will_update_count=existing_measurement_count,  # 若选 overwrite 则全部可更新
        unknown_point_count=unknown_point_count,
        file_duplicate_count=file_duplicate_count,
        warnings=all_warnings,
        errors=[e for e in errors if e.severity == "error"],
        items=items,
        can_confirm=can_confirm,
    )


@router.post("/api/projects/{project_id}/measurements/import-xlsx/preview", response_model=XlsxImportPreview)
async def preview_xlsx_import(
    project_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
) -> XlsxImportPreview:
    """第一步：上传 XLSX，解析并返回详细预检报告。

    不写入数据库。将解析结果和导入上下文临时保存，供 confirm 阶段使用。
    """
    project = db.get(models.Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="请上传 XLSX/XLSM 文件，旧版 .xls 格式暂不支持")

    file_bytes = await file.read()
    rows, errors = _parse_xlsx_rows_lenient(file_bytes)
    preview = _build_xlsx_preview(db, project_id, rows, errors, filename=file.filename or "unknown.xlsx")

    # 暂存解析结果和导入上下文到临时文件
    preview_id = f"XLSX-{uuid.uuid4().hex[:12]}"
    XLSX_TEMP_DIR.mkdir(parents=True, exist_ok=True)

    # 构建完整的临时数据（为 confirm 阶段准备）
    points = db.execute(
        select(models.TestPoint).where(models.TestPoint.project_db_id == project_id)
    ).scalars().all()
    point_lookup = {p.point_id: p.id for p in points}

    # 预计算 cycle_count → existing run_id 映射
    existing_runs = db.execute(
        select(models.TestRun).where(models.TestRun.project_db_id == project_id)
    ).scalars().all()
    cycle_to_run = {run.cycle_count: run.id for run in existing_runs}

    # 按 cycle_count 汇总第一条 run_name / test_time（用于新建 TestRun）
    cycle_meta: dict[int, dict] = {}
    for row in rows:
        cc = row.get("cycle_count", 0)
        if cc > 0 and cc not in cycle_meta:
            cycle_meta[cc] = {
                "run_name": row.get("run_name") or None,
                "test_time": row.get("test_time") or None,
            }

    temp_data = {
        "rows": rows,
        "project_id": project_id,
        "filename": file.filename or "unknown.xlsx",
        "items": [item.model_dump() for item in preview.items],
        "point_lookup": point_lookup,
        "cycle_to_run": {str(k): v for k, v in cycle_to_run.items()},
        "cycle_meta": {str(k): v for k, v in cycle_meta.items()},
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }

    temp_path = XLSX_TEMP_DIR / f"{preview_id}.json"
    temp_path.write_text(
        json.dumps(temp_data, ensure_ascii=False, default=str),
        encoding="utf-8",
    )

    preview.preview_id = preview_id
    return preview


@router.post("/api/projects/{project_id}/measurements/import-xlsx/confirm", response_model=XlsxImportResult)
def confirm_xlsx_import(
    project_id: int,
    payload: XlsxImportConfirmRequest,
    db: Session = Depends(get_db),
) -> XlsxImportResult:
    """第二步：确认导入，按用户选择的策略执行。

    从 preview 阶段保存的临时数据读取解析结果，
    不再重新解析 XLSX 文件。
    使用事务确保全部写入或全部回滚。
    """
    project = db.get(models.Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")

    # 从临时文件读取 preview 阶段保存的数据
    temp_path = XLSX_TEMP_DIR / f"{payload.preview_id}.json"
    if not temp_path.exists():
        raise HTTPException(status_code=404, detail="预检数据已过期，请重新上传预览")

    try:
        data = json.loads(temp_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError) as exc:
        raise HTTPException(status_code=400, detail=f"预检数据损坏: {exc}") from exc

    stored_project_id = data.get("project_id")
    if stored_project_id != project_id:
        raise HTTPException(status_code=400, detail="预检数据与当前项目不匹配")

    rows: list[dict] = data.get("rows", [])
    if not rows:
        raise HTTPException(status_code=400, detail="没有可导入的数据")

    point_lookup: dict[str, int] = data.get("point_lookup", {})
    cycle_to_run: dict[int, int] = {int(k): v for k, v in data.get("cycle_to_run", {}).items()}
    cycle_meta: dict[int, dict] = {int(k): v for k, v in data.get("cycle_meta", {}).items()}
    filename = data.get("filename", "unknown.xlsx")
    items_data: list[dict] = data.get("items", [])

    # 从 items 中重新构建状态分类
    item_status: dict[int, str] = {}  # row_index → status
    for item in items_data:
        item_status[item["row_index"]] = item["status"]

    # 加载点位（通过 point_lookup 中已验证的映射）
    points = db.execute(
        select(models.TestPoint).where(models.TestPoint.project_db_id == project_id)
    ).scalars().all()
    point_by_id: dict[str, models.TestPoint] = {p.point_id: p for p in points}

    strategy = payload.strategy
    skip_unknown = payload.skip_unknown_points
    skip_dupes = payload.skip_file_duplicates

    # ── strict 策略：前置校验 ──
    if strategy == XlsxImportStrategyEnum.STRICT:
        has_existing = any(s == XlsxRowStatus.EXISTING_MEASUREMENT for s in item_status.values())
        has_unknown = any(s == XlsxRowStatus.UNKNOWN_POINT for s in item_status.values())
        has_dupes = any(s == XlsxRowStatus.FILE_DUPLICATE for s in item_status.values())
        has_invalid = any(s == XlsxRowStatus.INVALID for s in item_status.values())
        if has_existing or has_unknown or has_dupes or has_invalid:
            blockers: list[str] = []
            if has_existing:
                blockers.append("存在已有测量记录")
            if has_unknown:
                blockers.append("存在未知点位")
            if has_dupes:
                blockers.append("存在文件内重复")
            if has_invalid:
                blockers.append("存在无效行")
            raise HTTPException(
                status_code=400,
                detail=f"严格模式下不允许导入：{'；'.join(blockers)}。请调整策略或修正 XLSX 后重试。"
            )

    # ── 统计计数器 ──
    created_runs = 0
    created_measurements = 0
    updated_measurements = 0
    filled_missing = 0
    skipped_existing = 0
    skipped_invalid = 0
    skipped_unknown = 0
    skipped_dupes = 0
    affected_point_ids: set[int] = set()

    # ── 按 cycle_count 分组的 run 缓存 ──
    run_cache: dict[int, models.TestRun] = {}

    def _get_or_create_run(cycle_count: int) -> tuple[models.TestRun, bool]:
        """按 project_db_id + cycle_count 查找或创建 TestRun。
        返回 (TestRun, is_newly_created)。
        """
        if cycle_count in run_cache:
            return run_cache[cycle_count], False

        # 查找已有 TestRun
        if cycle_count in cycle_to_run:
            run = db.get(models.TestRun, cycle_to_run[cycle_count])
            if run:
                run_cache[cycle_count] = run
                return run, False

        # 新建 TestRun
        meta = cycle_meta.get(cycle_count, {})
        run_name = meta.get("run_name") or f"{cycle_count}次"
        test_time = meta.get("test_time") or None
        run = models.TestRun(
            project_db_id=project_id,
            run_name=run_name,
            cycle_count=cycle_count,
            test_time=test_time,
            remark=f"Created by XLSX import: {filename}",
        )
        db.add(run)
        db.flush()
        run_cache[cycle_count] = run
        return run, True

    # ── 逐行处理 ──
    for row in rows:
        row_num = row["row_number"]
        status = item_status.get(row_num, XlsxRowStatus.INVALID)
        point_id = row.get("point_id") or ""
        cycle_count = row.get("cycle_count", 0)
        max_strain = row.get("max_strain_ue")
        min_strain = row.get("min_strain_ue")
        remark = row.get("remark")

        # 处理无效行
        if status == XlsxRowStatus.INVALID:
            skipped_invalid += 1
            continue

        # 处理未知点位
        if status == XlsxRowStatus.UNKNOWN_POINT:
            if skip_unknown:
                skipped_unknown += 1
                continue
            else:
                raise HTTPException(
                    status_code=400,
                    detail=f"第 {row_num} 行点位 {point_id} 不存在于当前项目。"
                           f"请先在点位管理中新增点位，或勾选「跳过未知点位」选项。"
                )

        # 处理文件内重复
        if status == XlsxRowStatus.FILE_DUPLICATE:
            if skip_dupes:
                skipped_dupes += 1
                continue
            else:
                raise HTTPException(
                    status_code=400,
                    detail=f"第 {row_num} 行在文件中重复 (cycle_count={cycle_count}, point_id={point_id})。"
                           f"请检查 XLSX 或勾选「跳过文件内重复」选项。"
                )

        # 获取或创建 TestRun
        run, is_new_run = _get_or_create_run(cycle_count)
        if is_new_run:
            created_runs += 1

        # 获取点位
        point = point_by_id.get(point_id) or db.get(models.TestPoint, point_lookup.get(point_id))
        if not point:
            if skip_unknown:
                skipped_unknown += 1
                continue
            raise HTTPException(status_code=400, detail=f"第 {row_num} 行点位 {point_id} 不存在")

        # 查找已有 MeasurementRecord
        existing = db.scalar(
            select(models.MeasurementRecord).where(
                models.MeasurementRecord.run_id == run.id,
                models.MeasurementRecord.point_db_id == point.id,
            )
        )

        # ── 按策略执行 ──
        if strategy == XlsxImportStrategyEnum.APPEND_ONLY:
            if existing:
                skipped_existing += 1
                continue
            record = models.MeasurementRecord(
                run_id=run.id,
                point_db_id=point.id,
                max_strain_ue=max_strain,
                min_strain_ue=min_strain,
                remark=remark,
            )
            _validate_strain_consistency(record.max_strain_ue, record.min_strain_ue)
            compute_measurement_fields(record)
            db.add(record)
            created_measurements += 1
            affected_point_ids.add(point.id)

        elif strategy == XlsxImportStrategyEnum.FILL_MISSING:
            if not existing:
                record = models.MeasurementRecord(
                    run_id=run.id,
                    point_db_id=point.id,
                    max_strain_ue=max_strain,
                    min_strain_ue=min_strain,
                    remark=remark,
                )
                _validate_strain_consistency(record.max_strain_ue, record.min_strain_ue)
                compute_measurement_fields(record)
                db.add(record)
                created_measurements += 1
            else:
                changed = False
                if existing.max_strain_ue is None and max_strain is not None:
                    existing.max_strain_ue = max_strain
                    changed = True
                if existing.min_strain_ue is None and min_strain is not None:
                    existing.min_strain_ue = min_strain
                    changed = True
                if existing.remark is None and remark is not None:
                    existing.remark = remark
                    changed = True
                if changed:
                    _validate_strain_consistency(existing.max_strain_ue, existing.min_strain_ue)
                    compute_measurement_fields(existing)
                    filled_missing += 1
                    affected_point_ids.add(point.id)
                else:
                    skipped_existing += 1

        elif strategy == XlsxImportStrategyEnum.OVERWRITE:
            if not existing:
                record = models.MeasurementRecord(
                    run_id=run.id,
                    point_db_id=point.id,
                    max_strain_ue=max_strain,
                    min_strain_ue=min_strain,
                    remark=remark,
                )
                _validate_strain_consistency(record.max_strain_ue, record.min_strain_ue)
                compute_measurement_fields(record)
                db.add(record)
                created_measurements += 1
            else:
                if max_strain is not None:
                    existing.max_strain_ue = max_strain
                if min_strain is not None:
                    existing.min_strain_ue = min_strain
                if remark is not None:
                    existing.remark = remark
                _validate_strain_consistency(existing.max_strain_ue, existing.min_strain_ue)
                compute_measurement_fields(existing)
                updated_measurements += 1
                affected_point_ids.add(point.id)

        elif strategy == XlsxImportStrategyEnum.STRICT:
            # strict 已经在前置校验通过，到这里都是新增
            record = models.MeasurementRecord(
                run_id=run.id,
                point_db_id=point.id,
                max_strain_ue=max_strain,
                min_strain_ue=min_strain,
                remark=remark,
            )
            _validate_strain_consistency(record.max_strain_ue, record.min_strain_ue)
            compute_measurement_fields(record)
            db.add(record)
            created_measurements += 1
            affected_point_ids.add(point.id)

    # ── 刷新异常标记并提交 ──
    db.flush()
    for pid in affected_point_ids:
        refresh_point_abnormal_flags(db, pid)
    db.commit()

    # 清理临时文件
    try:
        temp_path.unlink()
    except OSError:
        pass

    strategy_label = {
        XlsxImportStrategyEnum.APPEND_ONLY: "append_only",
        XlsxImportStrategyEnum.FILL_MISSING: "fill_missing",
        XlsxImportStrategyEnum.OVERWRITE: "overwrite",
        XlsxImportStrategyEnum.STRICT: "strict",
    }.get(strategy, strategy.value if hasattr(strategy, 'value') else str(strategy))

    return XlsxImportResult(
        success=True,
        strategy=strategy_label,
        created_run_count=created_runs,
        created_measurement_count=created_measurements,
        updated_measurement_count=updated_measurements,
        filled_missing_count=filled_missing,
        skipped_existing_count=skipped_existing,
        skipped_invalid_count=skipped_invalid,
        skipped_unknown_point_count=skipped_unknown,
        skipped_file_duplicate_count=skipped_dupes,
        message=f"导入完成：新增 {created_runs} 个轮次，{created_measurements} 条测量记录"
                + (f"，更新 {updated_measurements} 条" if updated_measurements else "")
                + (f"，填补 {filled_missing} 条" if filled_missing else "")
                + (f"，跳过 {skipped_existing} 条已有记录" if skipped_existing else ""),
    )
