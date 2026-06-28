import json
import re
import shutil
import tempfile
import uuid
from datetime import datetime
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app import models
from app.database import STORAGE_DIR


PROJECT_ROOT = Path(__file__).resolve().parents[2]
BACKUP_FILENAME = "pointprocess_backup.json"
RECORDS_XLSX = "records.xlsx"


def _iso(value: datetime | None) -> str | None:
    return value.isoformat() if value else None


def _safe_part(value: str | None, fallback: str) -> str:
    text = (value or fallback).strip() or fallback
    return re.sub(r"[^A-Za-z0-9._\-\u4e00-\u9fff]+", "_", text)


def _ext(filename: str | None, content_type: str | None = None) -> str:
    suffix = Path(filename or "").suffix
    if suffix:
        return suffix
    if content_type == "image/png":
        return ".png"
    if content_type in ("image/jpeg", "image/jpg"):
        return ".jpg"
    return ".bin"


def _rel_to_project(path: Path) -> str:
    return str(path.relative_to(PROJECT_ROOT))


def _source_path(stored_path: str) -> Path:
    return PROJECT_ROOT / stored_path


def _copy_if_exists(source: Path, target: Path) -> None:
    if not source.exists() or not source.is_file():
        return
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, target)


def _point_json(point: models.TestPoint, media_paths: dict[int, str]) -> dict:
    channel = point.channels[0] if point.channels else None
    cae = point.cae_mappings[0] if point.cae_mappings else None
    return {
        "id": point.id,
        "point_id": point.point_id,
        "point_name": point.point_name,
        "point_type": point.point_type,
        "component": point.component,
        "side": point.side,
        "position_description": point.position_description,
        "direction": point.direction,
        "bridge_type": point.bridge_type,
        "resistance_ohm": point.resistance_ohm,
        "install_status": point.install_status,
        "check_status": point.check_status,
        "remark": point.remark,
        "raw_json": point.raw_json,
        "created_at": _iso(point.created_at),
        "updated_at": _iso(point.updated_at),
        "channel": {
            "device": channel.device,
            "channel_name": channel.channel_name,
            "unit": channel.unit,
            "sample_rate_hz": channel.sample_rate_hz,
            "remark": channel.remark,
        } if channel else None,
        "cae_mapping": {
            "cae_point_id": cae.cae_point_id,
            "cae_component": cae.cae_component,
            "cae_result_type": cae.cae_result_type,
            "danger_level": cae.danger_level,
            "remark": cae.remark,
        } if cae else None,
        "photos": [
            {
                "id": media.id,
                "photo_id": media.photo_id,
                "type": media.type,
                "path": media_paths[media.id],
                "original_path": media.path,
                "filename": Path(media_paths[media.id]).name,
                "original_filename": media.filename,
                "taken_time": media.taken_time,
                "sha256": media.sha256,
                "remark": media.remark,
            }
            for media in sorted(point.media_files, key=lambda item: (item.type, item.id))
            if media.id in media_paths
        ],
    }


def _manifest_point(point: models.TestPoint, media_paths: dict[int, str]) -> dict:
    point_data = _point_json(point, media_paths)
    return {
        "point_id": point_data["point_id"],
        "point_name": point_data["point_name"],
        "point_type": point_data["point_type"] or "strain",
        "component": point_data["component"],
        "side": point_data["side"],
        "position_description": point_data["position_description"],
        "direction": point_data["direction"],
        "bridge_type": point_data["bridge_type"],
        "resistance_ohm": point_data["resistance_ohm"],
        "install_status": point_data["install_status"] or "planned",
        "check_status": point_data["check_status"],
        "channel": point_data["channel"],
        "cae_mapping": point_data["cae_mapping"],
        "photos": [
            {
                "photo_id": photo["photo_id"] or f"PHOTO-{point.point_id}-{index + 1:03d}",
                "type": photo["type"],
                "path": photo["path"],
                "filename": photo["filename"],
                "taken_time": photo["taken_time"],
                "sha256": photo["sha256"],
                "remark": photo["remark"],
            }
            for index, photo in enumerate(point_data["photos"])
        ],
        "tags": [],
        "remark": point_data["remark"],
        "created_time": point_data["created_at"],
        "updated_time": point_data["updated_at"],
        "custom_fields": None,
    }


def _write_workbook(project: models.Project, points: list[models.TestPoint], runs: list[models.TestRun], records_by_run_point: dict[tuple[int, int], models.MeasurementRecord], cracks: list[models.CrackRecord], photo_paths: dict[int, str], crack_paths: dict[int, str], target: Path) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "项目概览"
    summary.append(["字段", "值"])
    summary.append(["项目ID", project.project_id])
    summary.append(["项目名称", project.project_name])
    summary.append(["测试对象", project.test_object])
    summary.append(["试验类型", project.test_type])
    summary.append(["部门", project.department])
    summary.append(["产品/车型", project.vehicle_or_product])
    summary.append(["试验阶段", project.test_stage])
    summary.append(["点位数量", len(points)])
    summary.append(["测试轮次数量", len(runs)])
    summary.append(["裂缝记录数量", len(cracks)])
    summary.append(["导出时间", datetime.utcnow().isoformat()])

    point_sheet = wb.create_sheet("点位清单")
    point_sheet.append(["点位编号", "点位名称", "类型", "部件", "方位", "位置描述", "方向", "桥路", "电阻", "安装状态", "检查状态", "备注"])
    for point in points:
        point_sheet.append([point.point_id, point.point_name, point.point_type, point.component, point.side, point.position_description, point.direction, point.bridge_type, point.resistance_ohm, point.install_status, point.check_status, point.remark])

    photo_sheet = wb.create_sheet("点位照片")
    photo_sheet.append(["点位编号", "点位名称", "照片类型", "导出路径", "原文件名", "备注"])
    for point in points:
        for media in sorted(point.media_files, key=lambda item: item.id):
            photo_sheet.append([point.point_id, point.point_name, media.type, photo_paths.get(media.id, ""), media.filename, media.remark])

    crack_sheet = wb.create_sheet("裂缝照片")
    crack_sheet.append(["点位编号", "点位名称", "循环次数", "轮次", "导出路径", "原文件名", "备注", "记录时间"])
    for crack in cracks:
        crack_sheet.append([crack.point.point_id, crack.point.point_name, crack.cycle_count, crack.run.run_name if crack.run else "", crack_paths.get(crack.id, ""), crack.filename, crack.remark, _iso(crack.created_at)])

    used_titles: set[str] = set(wb.sheetnames)
    for run in runs:
        base_title = _safe_part(str(run.cycle_count), f"run_{run.id}")[:31] or f"run_{run.id}"
        title = base_title
        index = 2
        while title in used_titles:
            suffix = f"_{index}"
            title = f"{base_title[:31 - len(suffix)]}{suffix}"
            index += 1
        used_titles.add(title)
        sheet = wb.create_sheet(title)
        sheet.append(["点位编号", "点位名称", "部件", "方向", "最大应变(ue)", "最小应变(ue)", "应变幅(ue)", "应变范围(ue)", "应力幅(MPa)", "是否异常", "异常原因", "备注"])
        for point in points:
            record = records_by_run_point.get((run.id, point.id))
            sheet.append([
                point.point_id,
                point.point_name,
                point.component,
                point.direction,
                record.max_strain_ue if record else None,
                record.min_strain_ue if record else None,
                record.amplitude_strain_ue if record else None,
                record.range_strain_ue if record else None,
                record.stress_amplitude_mpa if record else None,
                record.is_abnormal if record else None,
                record.abnormal_reason if record else None,
                record.remark if record else None,
            ])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDEFEA")
            cell.alignment = Alignment(horizontal="center")
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 36)
    wb.save(target)


def build_project_export_zip(db: Session, project_id: int) -> tuple[Path, str]:
    project = db.execute(
        select(models.Project)
        .options(
            selectinload(models.Project.points).selectinload(models.TestPoint.media_files),
            selectinload(models.Project.points).selectinload(models.TestPoint.channels),
            selectinload(models.Project.points).selectinload(models.TestPoint.cae_mappings),
            selectinload(models.Project.test_runs).selectinload(models.TestRun.measurements),
            selectinload(models.Project.crack_records).selectinload(models.CrackRecord.point),
            selectinload(models.Project.crack_records).selectinload(models.CrackRecord.run),
            selectinload(models.Project.dewesoft_imports).selectinload(models.DewesoftImport.channels),
        )
        .where(models.Project.id == project_id)
    ).scalar_one_or_none()
    if not project:
        raise ValueError("project not found")

    points = sorted(project.points, key=lambda item: item.point_id)
    runs = sorted(project.test_runs, key=lambda item: (item.cycle_count, item.id))
    cracks = sorted(project.crack_records, key=lambda item: (item.cycle_count, item.point.point_id, item.id))
    dewesoft_imports = sorted(project.dewesoft_imports, key=lambda item: item.id)
    point_code_by_id = {point.id: point.point_id for point in points}
    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    export_id = f"WEB-{project.project_id}-{timestamp}-{uuid.uuid4().hex[:6]}"
    temp_dir = Path(tempfile.mkdtemp(prefix=f"pointprocess-export-{project.id}-", dir=STORAGE_DIR / "temp"))
    package_root = temp_dir / "package"
    package_root.mkdir(parents=True, exist_ok=True)

    photo_paths: dict[int, str] = {}
    photo_counters: dict[str, int] = {}
    for point in points:
        for media in sorted(point.media_files, key=lambda item: (item.type, item.id)):
            key = point.point_id
            photo_counters[key] = photo_counters.get(key, 0) + 1
            filename = f"{_safe_part(point.point_id, 'POINT')}_{media.type}_{photo_counters[key]:03d}{_ext(media.filename)}"
            relative = f"photos/{_safe_part(point.point_id, 'POINT')}/{filename}"
            _copy_if_exists(_source_path(media.stored_path), package_root / relative)
            photo_paths[media.id] = relative

    crack_paths: dict[int, str] = {}
    crack_counters: dict[str, int] = {}
    for crack in cracks:
        key = crack.point.point_id
        crack_counters[key] = crack_counters.get(key, 0) + 1
        filename = f"{_safe_part(crack.point.point_id, 'POINT')}_cycle_{crack.cycle_count}_{crack_counters[key]:03d}{_ext(crack.filename, crack.content_type)}"
        relative = f"cracks/{_safe_part(crack.point.point_id, 'POINT')}/{filename}"
        _copy_if_exists(_source_path(crack.stored_path), package_root / relative)
        crack_paths[crack.id] = relative

    records_by_run_point = {(record.run_id, record.point_db_id): record for run in runs for record in run.measurements}
    workbook_path = package_root / RECORDS_XLSX
    _write_workbook(project, points, runs, records_by_run_point, cracks, photo_paths, crack_paths, workbook_path)
    project_storage_root = STORAGE_DIR / "projects" / project.project_id
    for folder in ["raw", "attachments"]:
        source = project_storage_root / folder
        target = package_root / folder
        if source.exists() and source.is_dir():
            shutil.copytree(source, target, dirs_exist_ok=True)

    dewesoft_paths: dict[int, str] = {}
    for import_job in dewesoft_imports:
        source = _source_path(import_job.stored_path)
        relative = f"dewesoft/{import_job.id:04d}_{_safe_part(import_job.filename, 'dewesoft')}"
        _copy_if_exists(source, package_root / relative)
        dewesoft_paths[import_job.id] = relative

    manifest = {
        "schema_version": "1.0.0",
        "export_info": {
            "export_id": export_id,
            "export_time": datetime.utcnow().isoformat(),
            "app_name": "PointProcess Web",
            "app_version": "1.0",
            "remark": "PointProcess project export. manifest.json preserves app-compatible point and photo data; pointprocess_backup.json preserves full migration data.",
        },
        "project": {
            "project_id": project.project_id,
            "project_name": project.project_name,
            "test_object": project.test_object,
            "test_type": project.test_type,
            "department": project.department,
            "vehicle_or_product": project.vehicle_or_product,
            "test_stage": project.test_stage,
            "description": project.description,
            "created_time": _iso(project.created_at),
            "updated_time": _iso(project.updated_at),
        },
        "points": [_manifest_point(point, photo_paths) for point in points],
        "files": [
            {
                "file_id": "POINTPROCESS-RECORDS-XLSX",
                "type": "analysis_workbook",
                "path": RECORDS_XLSX,
                "filename": RECORDS_XLSX,
                "sha256": None,
                "remark": "Human-readable workbook generated by PointProcess.",
            }
        ],
        "custom_fields": {"pointprocess_backup": BACKUP_FILENAME},
    }
    (package_root / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    backup = {
        "format": "pointprocess_project_backup",
        "version": "1.0",
        "export_id": export_id,
        "exported_at": datetime.utcnow().isoformat(),
        "project": {
            "id": project.id,
            "project_id": project.project_id,
            "project_name": project.project_name,
            "test_object": project.test_object,
            "test_type": project.test_type,
            "department": project.department,
            "vehicle_or_product": project.vehicle_or_product,
            "test_stage": project.test_stage,
            "description": project.description,
            "source_export_id": project.source_export_id,
            "source_export_time": project.source_export_time,
            "raw_manifest_json": project.raw_manifest_json,
            "created_at": _iso(project.created_at),
            "updated_at": _iso(project.updated_at),
        },
        "points": [_point_json(point, photo_paths) for point in points],
        "test_runs": [
            {
                "id": run.id,
                "run_name": run.run_name,
                "cycle_count": run.cycle_count,
                "test_time": run.test_time,
                "remark": run.remark,
                "created_at": _iso(run.created_at),
                "measurements": [
                    {
                        "id": record.id,
                        "point_id": record.point.point_id,
                        "max_strain_ue": record.max_strain_ue,
                        "min_strain_ue": record.min_strain_ue,
                        "mean_strain_ue": record.mean_strain_ue,
                        "amplitude_strain_ue": record.amplitude_strain_ue,
                        "range_strain_ue": record.range_strain_ue,
                        "stress_max_mpa": record.stress_max_mpa,
                        "stress_min_mpa": record.stress_min_mpa,
                        "stress_mean_mpa": record.stress_mean_mpa,
                        "stress_amplitude_mpa": record.stress_amplitude_mpa,
                        "stress_range_mpa": record.stress_range_mpa,
                        "is_abnormal": record.is_abnormal,
                        "abnormal_reason": record.abnormal_reason,
                        "remark": record.remark,
                        "created_at": _iso(record.created_at),
                        "updated_at": _iso(record.updated_at),
                    }
                    for record in sorted(run.measurements, key=lambda item: item.point.point_id)
                ],
            }
            for run in runs
        ],
        "crack_records": [
            {
                "id": crack.id,
                "point_id": crack.point.point_id,
                "test_run_id": crack.test_run_id,
                "run_cycle_count": crack.run.cycle_count if crack.run else None,
                "cycle_count": crack.cycle_count,
                "path": crack_paths.get(crack.id),
                "filename": Path(crack_paths.get(crack.id, crack.filename)).name,
                "original_filename": crack.filename,
                "content_type": crack.content_type,
                "sha256": crack.sha256,
                "remark": crack.remark,
                "created_at": _iso(crack.created_at),
                "updated_at": _iso(crack.updated_at),
            }
            for crack in cracks
        ],
        "dewesoft_imports": [
            {
                "id": import_job.id,
                "test_run_id": import_job.test_run_id,
                "cycle_count": import_job.cycle_count,
                "run_name": import_job.run_name,
                "filename": import_job.filename,
                "path": dewesoft_paths.get(import_job.id),
                "status": import_job.status,
                "message": import_job.message,
                "duration_seconds": import_job.duration_seconds,
                "stable_start_seconds": import_job.stable_start_seconds,
                "stable_end_seconds": import_job.stable_end_seconds,
                "matched_channel_count": import_job.matched_channel_count,
                "unmatched_channel_count": import_job.unmatched_channel_count,
                "raw_metadata_json": import_job.raw_metadata_json,
                "created_at": _iso(import_job.created_at),
                "channels": [
                    {
                        "id": channel.id,
                        "channel_name": channel.channel_name,
                        "unit": channel.unit,
                        "sample_count": channel.sample_count,
                        "matched_point_id": point_code_by_id.get(channel.matched_point_db_id),
                        "measurement_id": channel.measurement_id,
                        "stable_min_strain_ue": channel.stable_min_strain_ue,
                        "stable_max_strain_ue": channel.stable_max_strain_ue,
                        "stable_mean_strain_ue": channel.stable_mean_strain_ue,
                        "raw_json": channel.raw_json,
                        "created_at": _iso(channel.created_at),
                    }
                    for channel in import_job.channels
                ],
            }
            for import_job in dewesoft_imports
        ],
    }
    (package_root / BACKUP_FILENAME).write_text(json.dumps(backup, ensure_ascii=False, indent=2), encoding="utf-8")

    zip_name = f"{_safe_part(project.project_id, 'project')}_pointprocess_export_{timestamp}.zip"
    zip_path = temp_dir / zip_name
    with ZipFile(zip_path, "w", ZIP_DEFLATED) as archive:
        for path in package_root.rglob("*"):
            if path.is_file():
                archive.write(path, path.relative_to(package_root).as_posix())
    return zip_path, zip_name
