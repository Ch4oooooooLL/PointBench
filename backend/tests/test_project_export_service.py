from openpyxl import load_workbook

from app import models
from app.services.project_export_service import _write_workbook


def test_workbook_includes_stress_change_sheet_before_cycle_sheets(tmp_path) -> None:
    project = models.Project(
        id=1,
        project_id="EXPORT-001",
        project_name="Export Test",
        raw_manifest_json="{}",
    )
    point_1 = models.TestPoint(
        id=1,
        project=project,
        point_id="P001",
        point_name="Point 001",
        point_type="strain",
        component="Frame",
        direction="X",
        install_status="installed",
        raw_json="{}",
    )
    point_2 = models.TestPoint(
        id=2,
        project=project,
        point_id="P002",
        point_name="Point 002",
        point_type="strain",
        component="Frame",
        direction="Y",
        install_status="installed",
        raw_json="{}",
    )
    run_1 = models.TestRun(id=11, project=project, run_name="Run 100", cycle_count=100)
    run_2 = models.TestRun(id=12, project=project, run_name="Run 200", cycle_count=200)
    records_by_run_point = {
        (run_1.id, point_1.id): models.MeasurementRecord(
            run=run_1,
            point=point_1,
            stress_amplitude_mpa=12.3,
        ),
        (run_2.id, point_1.id): models.MeasurementRecord(
            run=run_2,
            point=point_1,
            stress_amplitude_mpa=18.6,
        ),
        (run_1.id, point_2.id): models.MeasurementRecord(
            run=run_1,
            point=point_2,
            stress_amplitude_mpa=8.4,
        ),
    }
    target = tmp_path / "records.xlsx"

    _write_workbook(
        project,
        [point_1, point_2],
        [run_1, run_2],
        records_by_run_point,
        [],
        {},
        {},
        target,
    )

    workbook = load_workbook(target, data_only=True)

    assert workbook.sheetnames[:7] == [
        "项目概览",
        "点位清单",
        "点位照片",
        "裂缝照片",
        "应力变化情况",
        "100",
        "200",
    ]

    sheet = workbook["应力变化情况"]
    assert [cell.value for cell in sheet[1]] == [
        "点位编号",
        "点位名称",
        "部件",
        "方向",
        "100次应力幅(MPa)",
        "200次应力幅(MPa)",
    ]
    assert [cell.value for cell in sheet[2]] == ["P001", "Point 001", "Frame", "X", 12.3, 18.6]
    assert [cell.value for cell in sheet[3]] == ["P002", "Point 002", "Frame", "Y", 8.4, None]
